from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
import io
import os
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import base64
import json
import zipfile
import tempfile
from main import client, parse_json, parse_date, days_diff, ask_claude, pdf_to_content
from main import DARK_BLUE, MID_BLUE, LIGHT_BLUE, WHITE, GREEN, ORANGE

router = APIRouter()


@router.post("/reporte/pami")
async def reporte_pami(
    anio: str = Form(...), mes: str = Form(...), quincena: str = Form(...),
    caratula: UploadFile = File(...), opf: UploadFile = File(...),
    pre: UploadFile = File(...), pago: UploadFile = File(...),
    nr: UploadFile = File(...)
):
    car_bytes = await caratula.read()
    opf_bytes = await opf.read()
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()
    nr_bytes = await nr.read()

    periodo = f'{quincena} mes {mes} año {anio}'

    car_data = parse_json(ask_claude(
        pdf_to_content(car_bytes, 'CARÁTULA PAMI') + [{"type":"text","text":f"Período: {periodo}. Extraé: {{\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"total_pvp\":0.0,\"total_pvp_pami\":0.0,\"importe_bruto_convenio\":0.0}}"}],
        SYSTEM_JSON))

    opf_data = parse_json(ask_claude(
        pdf_to_content(opf_bytes, 'OPF PAMI') + [{"type":"text","text":f"Período: {periodo}. Buscar línea Efvo.PAMI. Extraé: {{\"efvo_pami\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":0}}"}],
        SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE PAMI') + [{"type":"text","text":"Extraé: {\"deb_cred_os\":0.0,\"bonif_tiras\":0.0,\"bonif_ambulatorio\":0.0,\"bonif_insulinas\":0.0,\"ret_gtos_adm_cofa\":0.0,\"efectivo_drogueria\":0.0,\"fdo_prest_colfarma\":0.0,\"nota_cred_ambulatorio\":0.0,\"nota_cred_insulina\":0.0,\"nota_cred_tiras\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. IMPORTANTE: deb_cred_os debe ser NEGATIVO si es un débito (el PRE lo muestra con signo negativo), positivo si es crédito, 0 si no existe. Todos los demás valores deben ser positivos (sin signo negativo). efectivo_drogueria = valor absoluto de EFECTIVO DROGUERIA SALDO."}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL PAMI') + [{"type":"text","text":"Buscar línea PAMI. Extraé: {\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}"}],
        SYSTEM_JSON))

    nr_text = xls_to_text(nr_bytes, nr.filename)
    nr_data = parse_json(ask_claude(
        [{"type":"text","text":f"NOTAS DE RECUPERO PAMI:\n{nr_text}\n\nLa tabla tiene columnas: TipoCte e ImporteCpte (entre otras). Sumá TODOS los valores de ImporteCpte agrupando por TipoCte (CCF, CCFD, NAF, NRFD, EfSa). Usá la columna ImporteCpte, NO ImporteCo. Fecha de Impresion de primera fila CCF o NAF = fecha_nr. Fecha de primera fila EfSa = fecha_efsa. Extraé: {{\"nr_ccf\":0.0,\"nr_ccfd\":0.0,\"nr_naf\":0.0,\"nr_nrfd\":0.0,\"nr_efsa\":0.0,\"fecha_nr\":\"DD/MM/YYYY\",\"fecha_efsa\":\"DD/MM/YYYY\"}}"}],
        SYSTEM_JSON))

    buf = build_pami_excel(
        {'caratula': car_data, 'opf': opf_data, 'pre': pre_data, 'pago': pago_data, 'nr': nr_data},
        quincena, mes, anio[-2:]
    )
    filename = f"{anio[-2:]}.{mes}.{quincena} - Reporte.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@router.post("/reporte/ioma")
async def reporte_ioma(
    anio: str = Form(...), mes: str = Form(...),
    agudo: UploadFile = File(...), opf: UploadFile = File(...),
    pre: UploadFile = File(...), pago: UploadFile = File(...),
    nr: UploadFile = File(...),
    planes: list[UploadFile] = File(default=[])
):
    agudo_bytes = await agudo.read()
    opf_bytes = await opf.read()
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()
    nr_bytes = await nr.read()

    agudo_data = parse_json(ask_claude(
        pdf_to_content(agudo_bytes, 'AGUDO/CRÓNICO IOMA') + [{"type":"text","text":"Buscar RX ON LINE/TOTALIDAD DE LAS RECETAS. Si es un Resumen de Facturación del Colegio, la fecha_cierre es la \"Fecha de Proceso\". Si es una Carátula On-Line, la fecha_cierre es la \"Fecha de generación\". Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}"}],
        SYSTEM_JSON))

    planes_data = {
        'AGUDO/CRÓNICO': {'recetas': agudo_data['recetas'], 'importe100': agudo_data['importe100'], 'ac_instituto': agudo_data['ac_instituto']},
        'RECURSOS DE AMPARO': {'recetas':0,'importe100':0,'ac_instituto':0},
        'RESOLUCIÓN DE DIRECTORIO': {'recetas':0,'importe100':0,'ac_instituto':0},
        'MAMI': {'recetas':0,'importe100':0,'ac_instituto':0},
        'MAYOR COBERTURA': {'recetas':0,'importe100':0,'ac_instituto':0},
        'VACUNAS': {'recetas':0,'importe100':0,'ac_instituto':0},
    }

    for plan_file in planes:
        pb = await plan_file.read()
        # Detectar si es Resumen del Colegio (múltiples planes) o carátula individual
        tipo_data = parse_json(ask_claude(
            pdf_to_content(pb, 'DOCUMENTO PLAN IOMA') + [{"type":"text","text":"Este documento es un Resumen de Facturación del Colegio de Farmacéuticos con tabla de múltiples planes, O es una carátula individual de un solo plan IOMA. Respondé SOLO: {\"tipo\":\"resumen_colegio\"} o {\"tipo\":\"caratula_individual\"}"}],
            SYSTEM_JSON))
        if tipo_data.get('tipo') == 'resumen_colegio':
            # Resumen del Colegio: extraer todos los planes de la tabla
            planes_lista = parse_json(ask_claude(
                pdf_to_content(pb, 'RESUMEN COLEGIO IOMA') + [{"type":"text","text":"Extraé cada fila de la tabla de planes. Para cada plan: importe100=Imp.Total, ac_instituto=Imp.Os. Respondé: {\"planes\":[{\"plan\":\"nombre del convenio/plan\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}]}"}],
                SYSTEM_JSON))
            for p in planes_lista.get('planes', []):
                plan_name = p.get('plan','').upper()
                entry = {'recetas':p['recetas'],'importe100':p['importe100'],'ac_instituto':p['ac_instituto']}
                if 'MAMI' in plan_name: planes_data['MAMI'] = entry
                elif 'MAYOR' in plan_name: planes_data['MAYOR COBERTURA'] = entry
                elif 'AMPARO' in plan_name: planes_data['RECURSOS DE AMPARO'] = entry
                elif 'DIRECTORIO' in plan_name: planes_data['RESOLUCIÓN DE DIRECTORIO'] = entry
                elif 'VACUNA' in plan_name: planes_data['VACUNAS'] = entry
        else:
            # Carátula individual de un solo plan
            pd_data = parse_json(ask_claude(
                pdf_to_content(pb, 'CARÁTULA PLAN IOMA') + [{"type":"text","text":"Leer Convenio/Plan para identificar el plan. Extraé: {\"plan\":\"nombre del plan\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}"}],
                SYSTEM_JSON))
            plan_name = pd_data.get('plan','').upper()
            entry = {'recetas':pd_data['recetas'],'importe100':pd_data['importe100'],'ac_instituto':pd_data['ac_instituto']}
            if 'MAMI' in plan_name: planes_data['MAMI'] = entry
            elif 'MAYOR' in plan_name: planes_data['MAYOR COBERTURA'] = entry
            elif 'AMPARO' in plan_name: planes_data['RECURSOS DE AMPARO'] = entry
            elif 'DIRECTORIO' in plan_name: planes_data['RESOLUCIÓN DE DIRECTORIO'] = entry
            elif 'VACUNA' in plan_name: planes_data['VACUNAS'] = entry

    opf_data = parse_json(ask_claude(
        pdf_to_content(opf_bytes, 'OPF IOMA') + [{"type":"text","text":"Efvo.IOMA AMBULATORIO = anticipo. Sección RETENCIONES línea RGI = ing_brutos_anticipo. Extraé: {\"efvo_ioma\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":0,\"ing_brutos_anticipo\":0.0}"}],
        SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE IOMA') + [{"type":"text","text":"Extraé: {\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"fdo_prest_colfarma\":0.0,\"nrf_ant\":0.0,\"nrf_def\":0.0,\"nrf_directas\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}"}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL IOMA') + [{"type":"text","text":"Sección RETENCIONES línea RGI = ing_brutos_pago. Extraé: {\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0,\"ing_brutos_pago\":0.0}"}],
        SYSTEM_JSON))

    nr_text = xls_to_text(nr_bytes, nr.filename)
    nr_data = parse_json(ask_claude(
        [{"type":"text","text":f"NOTAS DE RECUPERO IOMA:\n{nr_text}\n\nSumá NRF y NRFD agrupados por fecha de Impresion. Extraé: {{\"nr_por_fecha\":[{{\"fecha\":\"DD/MM/YYYY\",\"monto\":0.0}}]}}"}],
        SYSTEM_JSON))

    buf = build_ioma_excel(
        {'planes': planes_data, 'opf': opf_data, 'pre': pre_data, 'pago': pago_data, 'nr': nr_data, 'fecha_cierre': agudo_data['fecha_cierre']},
        mes, anio[-2:]
    )
    filename = f"{anio[-2:]}.{mes} - Reporte IOMA.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@router.post("/reporte/osde")
async def reporte_osde(
    anio: str = Form(...), mes: str = Form(...),
    caratula: UploadFile = File(...), pre: UploadFile = File(...),
    pago: UploadFile = File(...), nr: UploadFile = File(...)
):
    car_bytes = await caratula.read()
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()
    nr_bytes = await nr.read()

    car_data = parse_json(ask_claude(
        pdf_to_content(car_bytes, 'CARÁTULA OSDE') + [{"type":"text","text":"Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"afiliado\":0.0,\"a_cargo_osde\":0.0,\"bonificacion\":0.0,\"total_verificar\":0.0}"}],
        SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE OSDE') + [{"type":"text","text":"La tabla tiene columnas: Concepto, Base Cálculo, Créditos, Débitos. REGLAS: 1) Para retencion_fdo_res, ret_col_art12 y notas_credito tomá SIEMPRE el valor de la columna Débitos, NO la Base de Cálculo. 2) Para ajuste_facturacion: puede haber varias líneas de Ajuste Facturación. Identificá pares que se cancelan entre sí (mismo monto, uno en Débitos y otro en Créditos) y excluílos. El ajuste_facturacion es el monto de la línea que NO tiene contraparte que la cancele (si es débito es positivo, si es crédito es negativo). Si no hay ajuste real, ajuste_facturacion=0. 3) neto_cobrar = fila Neto a Cobrar columna Créditos. Extraé: {\"nro_liquidacion\":0,\"ajuste_facturacion\":0.0,\"retencion_fdo_res\":0.0,\"ret_col_art12\":0.0,\"notas_credito\":0.0,\"neto_cobrar\":0.0}"}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL OSDE') + [{"type":"text","text":f"La fecha_pago es la fecha del ENCABEZADO del documento (campo Fecha:), NO la fecha de presentación de la tabla. El nro_comprobante_pago es el número de Orden de Pago del encabezado. Confirmar que existe línea OSDE con liquidación nro {pre_data.get('nro_liquidacion','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}}"}],
        SYSTEM_JSON))

    nr_data = parse_json(ask_claude(
        pdf_to_content(nr_bytes, 'NOTA DE CRÉDITO DEL SUD') + [{"type":"text","text":"Tomar el TOTAL del documento. Extraé: {\"nr_monto\":0.0,\"nr_fecha\":\"DD/MM/YYYY\"}"}],
        SYSTEM_JSON))

    buf = build_osde_excel(
        {'caratula': car_data, 'pre': pre_data, 'pago': pago_data, 'nr': nr_data},
        mes, anio[-2:]
    )
    filename = f"{anio[-2:]}.{mes} - Reporte OSDE.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

# ── OSPECON ───────────────────────────────────────────────────────────────────

def build_ospecon_excel(data, mes, anio):
    car=data['caratula']; pre=data['pre']; pago=data['pago']
    fecha_pres=parse_date(car['fecha_cierre'])
    dias_pago=days_diff(fecha_pres,parse_date(pago['fecha_pago']))
    total_ret=abs(pre['retencion_fdo_res'])+abs(pre['ret_col_art12'])
    afiliado=car['importe_total']-car['ac_os']
    ajuste=pre.get('ajuste_facturacion',0)
    periodo=f'{mes}/{anio}'

    wb=Workbook(); ws=wb.active; ws.title='Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','OSPECON',bold=True,size=22,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',car['nro_recetas'],bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_pago),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')

    ws.merge_cells('B5:C5'); c(ws,'B5','IMPORTE TOTAL',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('B6:C6'); ni(ws,'B6',car['importe_total']); ws['B6'].font=Font(bold=True,size=13,color=WHITE); ws['B6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['B6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('E5:F5'); c(ws,'E5','A/C OSPECON',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('E6:F6'); ni(ws,'E6',car['ac_os']); ws['E6'].font=Font(bold=True,size=13,color=WHITE); ws['E6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['E6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('H5:I5'); c(ws,'H5','AFILIADO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('H6:I6'); ni(ws,'H6',afiliado); ws['H6'].font=Font(bold=True,size=13,color=WHITE); ws['H6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['H6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('K5:L5'); c(ws,'K5','TOTAL PAGADO OSPECON',size=9,color=WHITE,fill=GREEN,halign='center')
    ws.merge_cells('K6:L6'); ni(ws,'K6',pre['neto_cobrar']); ws['K6'].font=Font(bold=True,size=13,color=WHITE); ws['K6'].fill=PatternFill('solid',fgColor=GREEN); ws['K6'].alignment=Alignment(horizontal='center',vertical='center')

    c(ws,'B8','LIQ. FINAL',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    c(ws,'C8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ni(ws,'B9',pre['neto_cobrar']); ws['B9'].font=Font(bold=True,size=13,color=WHITE); ws['B9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['B9'].alignment=Alignment(horizontal='center',vertical='center')
    c(ws,'C9',dias_pago,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','PAGOS A FARMACIA',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','NETO A COBRAR',bold=True); n(ws,'C13',pre['neto_cobrar'])
    c(ws,'B14','Fecha pago'); d(ws,'C14',parse_date(pago['fecha_pago']))
    c(ws,'B15','Comprobante'); ws['C15'].value=pago['nro_comprobante_pago']; ws['C15'].alignment=Alignment(horizontal='right',vertical='center'); ws['C15'].font=Font(size=10)
    box(ws,11,2,15,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Total:'); n(ws,'F13',abs(pre['bonificacion']))
    ws.merge_cells('E14:F14'); c(ws,'E14','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E15','Fdo. Res.:'); n(ws,'F15',abs(pre['retencion_fdo_res']))
    c(ws,'E16','Colegio Art. 12 SU:'); n(ws,'F16',abs(pre['ret_col_art12']))
    c(ws,'E17','TOTAL',bold=True); n(ws,'F17',total_ret)
    ws.merge_cells('E18:F18'); c(ws,'E18','AJUSTE FACTURACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E19','Débito:'); n(ws,'F19',ajuste)
    box(ws,11,5,19,6)

    ws2=wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines=False
    style=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    h1=['RECETAS','IMPORTE TOTAL','A/C OSPECON','AFILIADO','%AFL','TOTAL PAGADO','%PAGADO','DIAS PAGO','RETENCIONES','%RET','BONIFICACIONES','%BON']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[car['nro_recetas'],car['importe_total'],car['ac_os'],afiliado,
          afiliado/car['importe_total'] if car['importe_total'] else 0,
          pre['neto_cobrar'],pre['neto_cobrar']/car['ac_os'] if car['ac_os'] else 0,
          dias_pago,total_ret,total_ret/car['ac_os'] if car['ac_os'] else 0,
          abs(pre['bonificacion']),abs(pre['bonificacion'])/car['ac_os'] if car['ac_os'] else 0]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_ospecon',ref='A1:L2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)
    for col in 'ABCDEFGHIJKL': ws2.column_dimensions[col].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@router.post("/reporte/ospecon")
async def reporte_ospecon(
    anio: str = Form(...), mes: str = Form(...),
    caratula: UploadFile = File(...), pre: UploadFile = File(...),
    pago: UploadFile = File(...)
):
    car_bytes = await caratula.read()
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()

    car_data = parse_json(ask_claude(
        pdf_to_content(car_bytes, 'CARÁTULA OSPECON') + [{"type":"text","text":"La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
        SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE OSPECON') + [{"type":"text","text":"Columnas: Concepto, Base Cálculo, Créditos, Débitos. Para bonificacion, retencion_fdo_res y ret_col_art12 tomá SIEMPRE la columna Débitos. Para ajuste_facturacion: puede haber varias líneas de Ajuste Facturación, identificá pares que se cancelan entre sí (mismo monto, uno Débitos y otro Créditos) y excluílos, el ajuste_facturacion es el monto de la línea que NO tiene contraparte (positivo si débito, negativo si crédito), 0 si no hay. neto_cobrar = fila Neto a Cobrar columna Créditos. Extraé: {\"nro_liquidacion\":0,\"ajuste_facturacion\":0.0,\"bonificacion\":0.0,\"retencion_fdo_res\":0.0,\"ret_col_art12\":0.0,\"neto_cobrar\":0.0}"}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL OSPECON') + [{"type":"text","text":f"La fecha_pago es la fecha del ENCABEZADO del documento (campo Fecha:). El nro_comprobante_pago es el número de Orden de Pago del encabezado. Confirmar línea OSPECON con liquidación nro {pre_data.get('nro_liquidacion','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}}"}],
        SYSTEM_JSON))

    buf = build_ospecon_excel(
        {'caratula': car_data, 'pre': pre_data, 'pago': pago_data},
        mes, anio[-2:]
    )
    filename = f"{anio[-2:]}.{mes} - Reporte OSPECON.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

# ── OSPRERA ───────────────────────────────────────────────────────────────────

def build_osprera_excel(data, mes, anio):
    planes = data['planes']; pre = data['pre']; pago = data['pago']
    opf = data.get('opf'); nr_data = data.get('nr')
    con_quincena = data.get('con_quincena', False)
    quincena = data.get('quincena', '')
    fecha_pres = parse_date(data.get('fecha_cierre') or pre['fecha_presentacion'])
    dias_pago = days_diff(fecha_pres, parse_date(pago['fecha_pago']))

    # Con quincena: puede haber OPF y NR
    dias_ant = 0; total_nr = 0; dias_nr_pond = 0
    if con_quincena and opf:
        dias_ant = days_diff(fecha_pres, parse_date(opf['fecha_opf']))
    if con_quincena and nr_data:
        nr_por_fecha = nr_data.get('nr_por_fecha', [])
        total_pond = 0
        for item in nr_por_fecha:
            m = item.get('monto', 0); total_nr += m
            total_pond += m * days_diff(fecha_pres, parse_date(item.get('fecha', '')))
        dias_nr_pond = total_pond / total_nr if total_nr else 0

    total_recetas = sum(p.get('recetas', 0) for p in planes.values())
    total_importe100 = sum(p.get('importe100', 0) for p in planes.values())
    total_ac = sum(p.get('ac_os', 0) for p in planes.values())
    afiliado = total_importe100 - total_ac

    deb_os = abs(pre['deb_cred_os']) if pre['deb_cred_os'] < 0 else 0
    cred_os = pre['deb_cred_os'] if pre['deb_cred_os'] > 0 else 0
    bonificaciones = abs(pre['bonificaciones'])
    ret_cofa = abs(pre['fdo_prest_colfarma']) + abs(pre['retencion_colegio_art12'])
    efvo_opf = opf['efvo_osprera'] if (con_quincena and opf) else 0
    liq_final = pre['total_liquidacion'] - efvo_opf
    total_pagado = efvo_opf + liq_final + total_nr
    if total_pagado == 0: total_pagado = pre['total_liquidacion']
    # dias prom ponderado si hay OPF o NR
    if con_quincena and (efvo_opf or total_nr):
        componentes = [(efvo_opf, dias_ant), (liq_final, dias_pago), (total_nr, dias_nr_pond)]
        total_pond2 = sum(v*d for v,d in componentes)
        dias_prom = total_pond2 / total_pagado if total_pagado else dias_pago
    else:
        dias_prom = dias_pago
    periodo = f'{mes}/{anio}' + (f' {quincena}' if con_quincena and quincena else '')
    planes_activos = {k: v for k, v in planes.items() if v.get('recetas', 0) > 0}

    wb = Workbook(); ws = wb.active; ws.title = 'Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','OSPRERA',bold=True,size=22,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',total_recetas,bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_prom,1),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0.0')

    for coord,label,val,clr in [('B5:C5','IMPORTE 100%',total_importe100,MID_BLUE),('E5:F5','A/C OSPRERA',total_ac,MID_BLUE),('H5:I5','AFILIADO',afiliado,MID_BLUE),('K5:L5','TOTAL PAGADO OSPRERA',total_pagado,GREEN)]:
        ws.merge_cells(coord); start=coord.split(':')[0]; end=coord.split(':')[1]
        r=int(start[1]); c_letter=start[0]; end_letter=end[0]
        c(ws,f'{c_letter}{r}',label,size=9,color=WHITE,fill=clr,halign='center')
        coord6=f'{c_letter}{r+1}:{end_letter}{r+1}'; ws.merge_cells(coord6)
        ni(ws,f'{c_letter}{r+1}',val); ws[f'{c_letter}{r+1}'].font=Font(bold=True,size=13,color=WHITE); ws[f'{c_letter}{r+1}'].fill=PatternFill('solid',fgColor=clr); ws[f'{c_letter}{r+1}'].alignment=Alignment(horizontal='center',vertical='center')

    if con_quincena and efvo_opf:
        for col,col_dias,lbl,dias,val in [('B','C','ANTICIPO',dias_ant,efvo_opf),('E','F','LIQ. FINAL',dias_pago,liq_final)]:
            c(ws,f'{col}8',lbl,size=9,color=WHITE,fill=MID_BLUE,halign='center')
            c(ws,f'{col_dias}8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
            ni(ws,f'{col}9',val); ws[f'{col}9'].font=Font(bold=True,size=13,color=WHITE); ws[f'{col}9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws[f'{col}9'].alignment=Alignment(horizontal='center',vertical='center')
            c(ws,f'{col_dias}9',dias,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')
        if total_nr:
            c(ws,'H8','NOTAS RECUP.',size=9,color=WHITE,fill=MID_BLUE,halign='center')
            c(ws,'I8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
            ni(ws,'H9',total_nr); ws['H9'].font=Font(bold=True,size=13,color=WHITE); ws['H9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['H9'].alignment=Alignment(horizontal='center',vertical='center')
            c(ws,'I9',round(dias_nr_pond,1),bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')
    else:
        c(ws,'B8','LIQ. FINAL',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        c(ws,'C8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        ni(ws,'B9',total_pagado); ws['B9'].font=Font(bold=True,size=13,color=WHITE); ws['B9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['B9'].alignment=Alignment(horizontal='center',vertical='center')
        c(ws,'C9',dias_pago,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','CARÁTULAS',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')
    ws.merge_cells('H11:L11'); c(ws,'H11','PAGOS A FARMACIA',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','COMPOSICIÓN POR PLAN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','PLAN',bold=True); c(ws,'C13','RECETAS',bold=True,halign='right')
    row = 14
    for plan, datos in planes_activos.items():
        c(ws,f'B{row}',plan); ws[f'C{row}'].value=datos['recetas']; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(size=10); row+=1
    c(ws,f'B{row}','TOTAL',bold=True); ws[f'C{row}'].value=total_recetas; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(bold=True,size=10)
    row_end_car=row; box(ws,11,2,row_end_car,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Total:'); n(ws,'F13',bonificaciones)
    ws.merge_cells('E14:F14'); c(ws,'E14','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E15','Fdo Prest. COLFARMA:'); n(ws,'F15',abs(pre['fdo_prest_colfarma']))
    c(ws,'E16','Colegio Art. 12 SU:'); n(ws,'F16',abs(pre['retencion_colegio_art12']))
    c(ws,'E17','TOTAL',bold=True); n(ws,'F17',ret_cofa)
    ws.merge_cells('E18:F18'); c(ws,'E18','DÉB. / CRÉD. OS',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E19','Débito OS:'); n(ws,'F19',deb_os)
    c(ws,'E20','Crédito OS:'); n(ws,'F20',cred_os)
    box(ws,11,5,20,6)

    if con_quincena and opf:
        ws.merge_cells('H12:L12'); c(ws,'H12','ANTICIPO (OPF)',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
        c(ws,'H13','TOTAL',bold=True); n(ws,'I13',efvo_opf)
        c(ws,'H14','Fecha pago'); d(ws,'I14',parse_date(opf['fecha_opf']))
        c(ws,'H15','Comprobante'); ws['I15'].value=opf['nro_comprobante_opf']; ws['I15'].alignment=Alignment(horizontal='right',vertical='center'); ws['I15'].font=Font(size=10)
        ws.merge_cells('H16:L16'); c(ws,'H16','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
        c(ws,'H17','Bruto a pagar antes imp.:'); n(ws,'I17',pre['total_liquidacion'])
        c(ws,'H18','TOTAL',bold=True); n(ws,'I18',liq_final)
        c(ws,'H19','Fecha pago'); d(ws,'I19',parse_date(pago['fecha_pago']))
        c(ws,'H20','Comprobante'); ws['I20'].value=pago['nro_comprobante_pago']; ws['I20'].alignment=Alignment(horizontal='right',vertical='center'); ws['I20'].font=Font(size=10)
        box_end = 20
        if total_nr and nr_data:
            nr_por_fecha = nr_data.get('nr_por_fecha', [])
            ws.merge_cells(f'H{box_end+1}:L{box_end+1}'); c(ws,f'H{box_end+1}','NOTAS DE RECUPERO',bold=True,size=10,fill=LIGHT_BLUE,halign='center'); box_end+=1
            c(ws,f'H{box_end+1}','Fecha',bold=True,halign='center'); c(ws,f'I{box_end+1}','Monto',bold=True,halign='center'); c(ws,f'K{box_end+1}','Días',bold=True,halign='center'); box_end+=1
            for item in sorted(nr_por_fecha, key=lambda x: x.get('fecha','')):
                df=days_diff(fecha_pres,parse_date(item.get('fecha',''))); d(ws,f'H{box_end+1}',parse_date(item.get('fecha',''))); n(ws,f'I{box_end+1}',item.get('monto',0))
                ws[f'K{box_end+1}'].value=df; ws[f'K{box_end+1}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'K{box_end+1}'].font=Font(size=10); box_end+=1
            c(ws,f'H{box_end+1}','DÍAS PROM.',bold=True); ws[f'K{box_end+1}'].value=round(dias_nr_pond,1); ws[f'K{box_end+1}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'K{box_end+1}'].font=Font(bold=True,size=10); box_end+=1
        box(ws,11,8,box_end,12)
    else:
        ws.merge_cells('H12:L12'); c(ws,'H12','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
        c(ws,'H13','TOTAL LIQUIDACIÓN',bold=True); n(ws,'I13',pre['total_liquidacion'])
        c(ws,'H14','Fecha pago'); d(ws,'I14',parse_date(pago['fecha_pago']))
        c(ws,'H15','Comprobante'); ws['I15'].value=pago['nro_comprobante_pago']; ws['I15'].alignment=Alignment(horizontal='right',vertical='center'); ws['I15'].font=Font(size=10)
        box(ws,11,8,15,12)

    ws2=wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines=False
    style=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    h1=['RECETAS','IMPORTE 100%','A/C OSPRERA','AFILIADO','%AFL','TOTAL PAGADO','%PAGADO','DIAS PAGO','DEBITOS OS','RETENCIONES','%RET','BONIFICACIONES','%BON']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[total_recetas,total_importe100,total_ac,afiliado,
          afiliado/total_importe100 if total_importe100 else 0,
          total_pagado,total_pagado/total_ac if total_ac else 0,
          dias_pago,deb_os,ret_cofa,ret_cofa/total_ac if total_ac else 0,
          bonificaciones,bonificaciones/total_ac if total_ac else 0]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_osprera',ref='A1:M2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)

    if con_quincena:
        ws2.cell(4,1,'Diferencias NR'); cell=ws2.cell(5,1); cell.value=dif_nr; cell.number_format='#,##0.00'
        tbl2=Table(displayName='tbl_osprera_dif',ref='A4:A5'); tbl2.tableStyleInfo=style; ws2.add_table(tbl2)

    for col in 'ABCDEFGHIJKLM': ws2.column_dimensions[col].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@router.post("/reporte/osprera")
async def reporte_osprera(
    anio: str = Form(...), mes: str = Form(...),
    caratulas: list[UploadFile] = File(...),
    pre: UploadFile = File(...),
    pago: UploadFile = File(...),
    quincena: str = Form(default=None),
    nr: UploadFile = File(default=None)
):
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()
    con_quincena = quincena is not None

    planes_data = {
        'GENERAL': {'recetas':0,'importe100':0.0,'ac_os':0.0},
        'TRATAMIENTO PROLONGADO': {'recetas':0,'importe100':0.0,'ac_os':0.0},
        'MONOTRIBUTISTAS': {'recetas':0,'importe100':0.0,'ac_os':0.0},
        'RURAL': {'recetas':0,'importe100':0.0,'ac_os':0.0},
        'DECLARACIÓN DE DISPENSA': {'recetas':0,'importe100':0.0,'ac_os':0.0},
    }

    fecha_cierre_osprera = None
    for car_file in caratulas:
        cb = await car_file.read()
        car_data = parse_json(ask_claude(
            pdf_to_content(cb, 'CARÁTULA OSPRERA') + [{"type":"text","text":"Identificar el plan exacto leyendo el campo Convenio/Plan. La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"plan\":\"nombre completo del convenio/plan\",\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
            SYSTEM_JSON))
        if not fecha_cierre_osprera: fecha_cierre_osprera = car_data.get('fecha_cierre')
        plan_name = car_data.get('plan','').upper()
        if 'MONOT' in plan_name: key='MONOTRIBUTISTAS'
        elif 'RURAL' in plan_name: key='RURAL'
        elif 'PROLONGADO' in plan_name: key='TRATAMIENTO PROLONGADO'
        elif 'DISPENSA' in plan_name: key='DECLARACIÓN DE DISPENSA'
        else: key='GENERAL'
        planes_data[key]['recetas'] += car_data.get('nro_recetas',0)
        planes_data[key]['importe100'] += car_data.get('importe_total',0.0)
        planes_data[key]['ac_os'] += car_data.get('ac_os',0.0)

    nr_data = None
    if con_quincena and nr:
        nr_bytes = await nr.read()
        nr_text = xls_to_text(nr_bytes, nr.filename)
        nr_data = parse_json(ask_claude(
            [{"type":"text","text":f"NOTAS DE RECUPERO OSPRERA:\n{nr_text}\n\nSumá NRF y NRFD agrupados por fecha. Extraé: {{\"nr_por_fecha\":[{{\"fecha\":\"DD/MM/YYYY\",\"monto\":0.0}}]}}"}],
            SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE OSPRERA') + [{"type":"text","text":"Extraé: {\"fecha_presentacion\":\"DD/MM/YYYY\",\"nro_comprobante\":0,\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"fdo_prest_colfarma\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. deb_cred_os = DEB/CRED DE OBRA SOCIAL (negativo si es débito). bonificaciones = BONIFICACIONES. fdo_prest_colfarma = FDO PREST COLFARMA. total_liquidacion = Total liquidación."}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL OSPRERA') + [{"type":"text","text":f"La fecha_pago es la Fecha del encabezado del documento. El nro_comprobante_pago es el número de Comprobante del encabezado. Confirmar que existe línea OSPRERA con comprobante {pre_data.get('nro_comprobante','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":\"\"}}"}],
        SYSTEM_JSON))

    buf = build_osprera_excel(
        {'planes': planes_data, 'pre': pre_data, 'pago': pago_data,
         'fecha_cierre': fecha_cierre_osprera, 'opf': None, 'nr': nr_data,
         'con_quincena': con_quincena, 'quincena': quincena},
        mes, anio[-2:]
    )
    q_str = f".{quincena}" if con_quincena else ""
    filename = f"{anio[-2:]}.{mes}{q_str} - Reporte OSPRERA.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

# ── UNION PERSONAL ─────────────────────────────────────────────────────────────

def build_unionpersonal_excel(data, mes, anio):
    planes = data['planes']; opf = data['opf']; pre = data['pre']; pago = data['pago']
    fecha_pres = parse_date(data.get('fecha_cierre') or pre['fecha_presentacion'])
    dias_ant = days_diff(fecha_pres, parse_date(opf['fecha_opf']))
    dias_liq = days_diff(fecha_pres, parse_date(pago['fecha_pago']))

    total_recetas = sum(p.get('recetas', 0) for p in planes.values())
    total_importe100 = sum(p.get('importe100', 0) for p in planes.values())
    total_ac = sum(p.get('ac_os', 0) for p in planes.values())
    afiliado = total_importe100 - total_ac

    bonificaciones = abs(pre['bonificaciones'])
    ret_cofa = abs(pre['fdo_prest_colfarma']) + abs(pre['retencion_colegio_art12'])
    deb_os = abs(pre.get('deb_cred_os', 0)) if pre.get('deb_cred_os', 0) < 0 else 0
    cred_os = pre.get('deb_cred_os', 0) if pre.get('deb_cred_os', 0) > 0 else 0
    liq_final = pre['total_liquidacion'] - opf['efvo_up']
    total_pagado = opf['efvo_up'] + liq_final
    dias_prom = (opf['efvo_up']*dias_ant + liq_final*dias_liq) / total_pagado if total_pagado else 0
    periodo = f'{mes}/{anio}'
    planes_activos = {k: v for k, v in planes.items() if v.get('recetas', 0) > 0}

    wb = Workbook(); ws = wb.active; ws.title = 'Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','UNIÓN PERSONAL',bold=True,size=16,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',total_recetas,bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_prom,1),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0.0')

    for coord,label,val,clr in [('B5:C5','IMPORTE 100%',total_importe100,MID_BLUE),('E5:F5','A/C UNIÓN PERSONAL',total_ac,MID_BLUE),('H5:I5','AFILIADO',afiliado,MID_BLUE),('K5:L5','TOTAL PAGADO UP',total_pagado,GREEN)]:
        ws.merge_cells(coord); start=coord.split(':')[0]; end=coord.split(':')[1]
        r=int(start[1]); c_letter=start[0]; end_letter=end[0]
        c(ws,f'{c_letter}{r}',label,size=9,color=WHITE,fill=clr,halign='center')
        coord6=f'{c_letter}{r+1}:{end_letter}{r+1}'; ws.merge_cells(coord6)
        ni(ws,f'{c_letter}{r+1}',val); ws[f'{c_letter}{r+1}'].font=Font(bold=True,size=13,color=WHITE); ws[f'{c_letter}{r+1}'].fill=PatternFill('solid',fgColor=clr); ws[f'{c_letter}{r+1}'].alignment=Alignment(horizontal='center',vertical='center')

    for col,col_dias,lbl,dias,val in [('B','C','ANTICIPO',dias_ant,opf['efvo_up']),('E','F','LIQ. FINAL',dias_liq,liq_final)]:
        c(ws,f'{col}8',lbl,size=9,color=WHITE,fill=MID_BLUE,halign='center')
        c(ws,f'{col_dias}8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        n(ws,f'{col}9',val); ws[f'{col}9'].font=Font(bold=True,size=13,color=WHITE); ws[f'{col}9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws[f'{col}9'].alignment=Alignment(horizontal='center',vertical='center')
        c(ws,f'{col_dias}9',dias,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','CARÁTULAS',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')
    ws.merge_cells('H11:L11'); c(ws,'H11','PAGOS A FARMACIA',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','COMPOSICIÓN POR PLAN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','PLAN',bold=True); c(ws,'C13','RECETAS',bold=True,halign='right')
    row = 14
    for plan, datos in planes_activos.items():
        c(ws,f'B{row}',plan); ws[f'C{row}'].value=datos['recetas']; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(size=10); row+=1
    c(ws,f'B{row}','TOTAL',bold=True); ws[f'C{row}'].value=total_recetas; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(bold=True,size=10)
    row_end_car=row; box(ws,11,2,row_end_car,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Total:'); n(ws,'F13',bonificaciones)
    ws.merge_cells('E14:F14'); c(ws,'E14','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E15','Fdo Prest. COLFARMA:'); n(ws,'F15',abs(pre['fdo_prest_colfarma']))
    c(ws,'E16','Colegio Art. 12 SU:'); n(ws,'F16',abs(pre['retencion_colegio_art12']))
    c(ws,'E17','TOTAL',bold=True); n(ws,'F17',ret_cofa)
    ws.merge_cells('E18:F18'); c(ws,'E18','DÉB. / CRÉD. OS',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E19','Débito OS:'); n(ws,'F19',deb_os)
    c(ws,'E20','Crédito OS:'); n(ws,'F20',cred_os)
    box(ws,11,5,20,6)

    ws.merge_cells('H12:L12'); c(ws,'H12','ANTICIPO (OPF)',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'H13','TOTAL',bold=True); n(ws,'I13',opf['efvo_up'])
    c(ws,'H14','Fecha pago'); d(ws,'I14',parse_date(opf['fecha_opf']))
    c(ws,'H15','Comprobante'); ws['I15'].value=opf['nro_comprobante_opf']; ws['I15'].alignment=Alignment(horizontal='right',vertical='center'); ws['I15'].font=Font(size=10)
    ws.merge_cells('H16:L16'); c(ws,'H16','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'H17','Bruto a pagar antes imp.:'); n(ws,'I17',pre['total_liquidacion'])
    c(ws,'H18','TOTAL',bold=True); n(ws,'I18',liq_final)
    c(ws,'H19','Fecha pago'); d(ws,'I19',parse_date(pago['fecha_pago']))
    c(ws,'H20','Comprobante'); ws['I20'].value=pago['nro_comprobante_pago']; ws['I20'].alignment=Alignment(horizontal='right',vertical='center'); ws['I20'].font=Font(size=10)
    box(ws,11,8,20,12)

    ws2=wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines=False
    style=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    h1=['RECETAS','IMPORTE 100%','A/C UP','AFILIADO','%AFL','TOTAL PAGADO','%PAGADO','DIAS PROM','RETENCIONES','%RET','BONIFICACIONES','%BON']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[total_recetas,total_importe100,total_ac,afiliado,
          afiliado/total_importe100 if total_importe100 else 0,
          total_pagado,total_pagado/total_ac if total_ac else 0,
          round(dias_prom,2),ret_cofa,ret_cofa/total_ac if total_ac else 0,
          bonificaciones,bonificaciones/total_ac if total_ac else 0]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_unionpersonal',ref='A1:L2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)
    h2=['ANTICIPO','%UP ANT','DIAS ANT','LIQ FINAL','%UP LIQ','DIAS LIQ']
    for i,h in enumerate(h2): ws2.cell(4,i+1,h)
    row5=[opf['efvo_up'],opf['efvo_up']/total_pagado if total_pagado else 0,dias_ant,
          liq_final,liq_final/total_pagado if total_pagado else 0,dias_liq]
    for i,v in enumerate(row5):
        cell=ws2.cell(5,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl2=Table(displayName='tbl_up_desglose',ref='A4:F5'); tbl2.tableStyleInfo=style; ws2.add_table(tbl2)
    for col in 'ABCDEFGHIJKL': ws2.column_dimensions[col].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@router.post("/reporte/unionpersonal")
async def reporte_unionpersonal(
    anio: str = Form(...), mes: str = Form(...),
    caratulas: list[UploadFile] = File(...),
    opf: UploadFile = File(...),
    pre: UploadFile = File(...),
    pago: UploadFile = File(...)
):
    opf_bytes = await opf.read()
    pre_bytes = await pre.read()
    pago_bytes = await pago.read()

    planes_data = {
        'PLANES VARIOS': {'recetas':0,'importe100':0.0,'ac_os':0.0},
        'DECLARACIÓN DE DISPENSA': {'recetas':0,'importe100':0.0,'ac_os':0.0},
    }

    fecha_cierre_up = None
    for car_file in caratulas:
        cb = await car_file.read()
        car_data = parse_json(ask_claude(
            pdf_to_content(cb, 'CARÁTULA UNIÓN PERSONAL') + [{"type":"text","text":"Identificar el plan: Planes Varios o Declaracion de dispensa. La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"plan\":\"nombre del plan\",\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
            SYSTEM_JSON))
        if not fecha_cierre_up: fecha_cierre_up = car_data.get('fecha_cierre')
        plan_name = car_data.get('plan','').upper()
        key = 'DECLARACIÓN DE DISPENSA' if 'DISPENSA' in plan_name else 'PLANES VARIOS'
        planes_data[key]['recetas'] += car_data.get('nro_recetas',0)
        planes_data[key]['importe100'] += car_data.get('importe_total',0.0)
        planes_data[key]['ac_os'] += car_data.get('ac_os',0.0)

    opf_data = parse_json(ask_claude(
        pdf_to_content(opf_bytes, 'OPF UNIÓN PERSONAL') + [{"type":"text","text":"Buscar línea UNION PERSONAL (SIFAR) con descripción que empiece con Efvo. La fecha_opf es la Fecha del encabezado. El nro_comprobante_opf es el Comprobante del encabezado. Extraé: {\"efvo_up\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":\"\"}"}],
        SYSTEM_JSON))

    pre_data = parse_json(ask_claude(
        pdf_to_content(pre_bytes, 'PRE UNIÓN PERSONAL') + [{"type":"text","text":"Extraé: {\"fecha_presentacion\":\"DD/MM/YYYY\",\"nro_comprobante\":0,\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"fdo_prest_colfarma\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. deb_cred_os = DEB/CRED DE OBRA SOCIAL (negativo si es débito, 0 si no aparece). bonificaciones=BONIFICACIONES, fdo_prest_colfarma=FDO PREST COLFARMA, total_liquidacion=Total liquidación."}],
        SYSTEM_JSON))

    pago_data = parse_json(ask_claude(
        pdf_to_content(pago_bytes, 'PAGO FINAL UNIÓN PERSONAL') + [{"type":"text","text":f"La fecha_pago es la Fecha del encabezado. El nro_comprobante_pago es el Comprobante del encabezado. Buscar línea UNION PERSONAL con liquidación nro {pre_data.get('nro_comprobante','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":\"\"}}"}],
        SYSTEM_JSON))

    buf = build_unionpersonal_excel(
        {'planes': planes_data, 'opf': opf_data, 'pre': pre_data, 'pago': pago_data, 'fecha_cierre': fecha_cierre_up},
        mes, anio[-2:]
    )
    filename = f"{anio[-2:]}.{mes} - Reporte Union Personal.xlsx"
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})

# ── BATCH ─────────────────────────────────────────────────────────────────────

import zipfile
import tempfile

def find_file(files, *keywords):
    """Busca un archivo que contenga alguna de las keywords en el nombre (case insensitive)"""
    for kw in keywords:
        for f in files:
            if kw.lower() in os.path.basename(f).lower():
                return f
    return None

def find_files(files, *keywords):
    """Busca todos los archivos que contengan alguna de las keywords"""
    result = []
    for kw in keywords:
        for f in files:
            if kw.lower() in os.path.basename(f).lower():
                if f not in result:
                    result.append(f)
    return result

def read_file(path):
    with open(path, 'rb') as f:
        return f.read()

@router.post("/batch/pami")
async def batch_pami(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        # Extraer ZIP
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []
        errores = []

        # Recorrer estructura: año/mes/quincena
        for root, dirs, files in os.walk(tmpdir):
            # Detectar si es carpeta de quincena (contiene 1Q o 2Q en el path)
            rel = os.path.relpath(root, tmpdir)
            parts = rel.replace('\\', '/').split('/')
            
            # Buscar nivel de quincena: tiene que tener archivos directamente
            if not any('1Q' in p or '2Q' in p for p in parts):
                continue
            
            all_files = [os.path.join(root, f) for f in os.listdir(root) if os.path.isfile(os.path.join(root, f))]
            
            # Buscar subcarpeta de liquidacion
            liq_dir = None
            for d in os.listdir(root):
                full_d = os.path.join(root, d)
                if os.path.isdir(full_d) and 'liquidac' in d.lower():
                    liq_dir = full_d
                    break
            
            if not liq_dir:
                continue

            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]
            
            caratula = find_file(all_files, 'caratula', 'carátula')
            nr_file = find_file(all_files, 'notas', 'recupero', '.xls')
            opf_file = find_file(liq_files, 'opf')
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')

            if not all([caratula, nr_file, opf_file, pre_file, pago_file]):
                errores.append(f"❌ Faltan archivos en: {rel}\n   → Verificá que la carpeta tenga todos los archivos necesarios (Carátula, PRE, Pago, etc.)")
                continue

            try:
                # Extraer quincena y mes/año del path
                quincena = next((p for p in parts if '1Q' in p or '2Q' in p), None)
                quincena_val = '1Q' if quincena and '1Q' in quincena else '2Q'
                
                # Buscar mes y año en el path
                mes_part = next((p for p in parts if any(m in p for m in ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'])), None)
                anio_part = next((p for p in parts if p.isdigit() and len(p) == 4), None)
                
                if not mes_part or not anio_part:
                    errores.append(f"❌ No se pudo determinar mes/año en: {rel}\n   → Verificá que la estructura de carpetas sea: año/mes/quincena")
                    continue

                meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}
                mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
                anio = anio_part

                car_bytes = read_file(caratula)
                opf_bytes = read_file(opf_file)
                pre_bytes = read_file(pre_file)
                pago_bytes = read_file(pago_file)
                nr_bytes = read_file(nr_file)
                nr_filename = os.path.basename(nr_file)

                periodo = f'{quincena_val} mes {mes_num} año {anio}'

                car_data = parse_json(ask_claude(
                    pdf_to_content(car_bytes, 'CARÁTULA PAMI') + [{"type":"text","text":f"Período: {periodo}. Extraé: {{\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"total_pvp\":0.0,\"total_pvp_pami\":0.0,\"importe_bruto_convenio\":0.0}}"}],
                    SYSTEM_JSON))
                opf_data = parse_json(ask_claude(
                    pdf_to_content(opf_bytes, 'OPF PAMI') + [{"type":"text","text":f"Período: {periodo}. Buscar línea Efvo.PAMI. Extraé: {{\"efvo_pami\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":0}}"}],
                    SYSTEM_JSON))
                pre_data = parse_json(ask_claude(
                    pdf_to_content(pre_bytes, 'PRE PAMI') + [{"type":"text","text":"Extraé: {\"deb_cred_os\":0.0,\"bonif_tiras\":0.0,\"bonif_ambulatorio\":0.0,\"bonif_insulinas\":0.0,\"ret_gtos_adm_cofa\":0.0,\"efectivo_drogueria\":0.0,\"fdo_prest_colfarma\":0.0,\"nota_cred_ambulatorio\":0.0,\"nota_cred_insulina\":0.0,\"nota_cred_tiras\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. IMPORTANTE: deb_cred_os debe ser NEGATIVO si es un débito (el PRE lo muestra con signo negativo), positivo si es crédito, 0 si no existe. Todos los demás valores deben ser positivos (sin signo negativo). efectivo_drogueria = valor absoluto de EFECTIVO DROGUERIA SALDO."}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(pago_bytes, 'PAGO FINAL PAMI') + [{"type":"text","text":"Buscar línea PAMI. Extraé: {\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}"}],
                    SYSTEM_JSON))
                nr_text = xls_to_text(nr_bytes, nr_filename)
                nr_data = parse_json(ask_claude(
                    [{"type":"text","text":f"NOTAS DE RECUPERO PAMI:\n{nr_text}\n\nLa tabla tiene columnas: TipoCte e ImporteCpte (entre otras). Sumá TODOS los valores de ImporteCpte agrupando por TipoCte (CCF, CCFD, NAF, NRFD, EfSa). Usá la columna ImporteCpte, NO ImporteCo. Fecha de Impresion de primera fila CCF o NAF = fecha_nr. Fecha de primera fila EfSa = fecha_efsa. Extraé: {{\"nr_ccf\":0.0,\"nr_ccfd\":0.0,\"nr_naf\":0.0,\"nr_nrfd\":0.0,\"nr_efsa\":0.0,\"fecha_nr\":\"DD/MM/YYYY\",\"fecha_efsa\":\"DD/MM/YYYY\"}}"}],
                    SYSTEM_JSON))

                buf = build_pami_excel(
                    {'caratula': car_data, 'opf': opf_data, 'pre': pre_data, 'pago': pago_data, 'nr': nr_data},
                    quincena_val, mes_num, anio[-2:]
                )
                filename = f"{anio[-2:]}.{mes_num}.{quincena_val} - Reporte.xlsx"
                reportes.append((filename, buf.getvalue()))

            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")

        # Crear ZIP de salida
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_PAMI.zip"'})

@router.post("/batch/ioma")
async def batch_ioma(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []; errores = []
        meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}

        for root, dirs, files in os.walk(tmpdir):
            rel = os.path.relpath(root, tmpdir).replace('\\','/')
            parts = rel.split('/')
            # Nivel de mes: contiene nombre de mes y tiene subcarpetas de caratulas/liquidaciones
            if not any(m in rel for m in meses_map.keys()): continue
            if not any(os.path.isdir(os.path.join(root, d)) for d in os.listdir(root)): continue

            all_files = [os.path.join(root, f) for f in os.listdir(root) if os.path.isfile(os.path.join(root, f))]
            subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]

            car_dir = next((d for d in subdirs if 'caratula' in os.path.basename(d).lower()), None)
            liq_dir = next((d for d in subdirs if 'liquidac' in os.path.basename(d).lower()), None)
            nr_file = find_file(all_files, 'notas', 'recupero', '.xls')

            if not liq_dir: continue

            car_files = sorted([os.path.join(car_dir, f) for f in os.listdir(car_dir)]) if car_dir else []
            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]

            opf_file = find_file(liq_files, 'opf')
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')

            if not all([car_files, opf_file, pre_file, pago_file]):
                errores.append(f"❌ Faltan archivos en: {rel}\n   → Verificá que la carpeta tenga todos los archivos necesarios (Carátula, PRE, Pago, etc.)"); continue

            anio_part = next((p for p in parts if p.isdigit() and len(p)==4), None)
            mes_part = next((p for p in parts if any(m in p for m in meses_map.keys())), None)
            if not anio_part or not mes_part:
                errores.append(f"❌ No se pudo determinar mes/año en: {rel}\n   → Verificá que la estructura de carpetas sea: año/mes/quincena"); continue

            mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
            anio = anio_part

            try:
                agudo_bytes = read_file(car_files[0])
                agudo_data = parse_json(ask_claude(
                    pdf_to_content(agudo_bytes, 'AGUDO/CRÓNICO IOMA') + [{"type":"text","text":"Buscar RX ON LINE/TOTALIDAD DE LAS RECETAS. Si es un Resumen de Facturación del Colegio, la fecha_cierre es la \"Fecha de Proceso\". Si es una Carátula On-Line, la fecha_cierre es la \"Fecha de generación\". Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}"}],
                    SYSTEM_JSON))

                planes_data = {
                    'AGUDO/CRÓNICO': {'recetas':agudo_data['recetas'],'importe100':agudo_data['importe100'],'ac_instituto':agudo_data['ac_instituto']},
                    'RECURSOS DE AMPARO': {'recetas':0,'importe100':0,'ac_instituto':0},
                    'RESOLUCIÓN DE DIRECTORIO': {'recetas':0,'importe100':0,'ac_instituto':0},
                    'MAMI': {'recetas':0,'importe100':0,'ac_instituto':0},
                    'MAYOR COBERTURA': {'recetas':0,'importe100':0,'ac_instituto':0},
                    'VACUNAS': {'recetas':0,'importe100':0,'ac_instituto':0},
                }

                for cb_path in car_files[1:]:
                    cb = read_file(cb_path)
                    # Detectar tipo
                    tipo_data = parse_json(ask_claude(
                        pdf_to_content(cb, 'DOCUMENTO PLAN IOMA') + [{"type":"text","text":"Este documento es un Resumen de Facturación del Colegio con tabla de múltiples planes, O es una carátula individual de un solo plan IOMA. Respondé SOLO: {\"tipo\":\"resumen_colegio\"} o {\"tipo\":\"caratula_individual\"}"}],
                        SYSTEM_JSON))
                    if tipo_data.get('tipo') == 'resumen_colegio':
                        planes_lista = parse_json(ask_claude(
                            pdf_to_content(cb, 'RESUMEN COLEGIO IOMA') + [{"type":"text","text":"Extraé cada fila de la tabla de planes. Para cada plan: importe100=Imp.Total, ac_instituto=Imp.Os. Respondé: {\"planes\":[{\"plan\":\"nombre del convenio/plan\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}]}"}],
                            SYSTEM_JSON))
                        for p in planes_lista.get('planes', []):
                            plan_name = p.get('plan','').upper()
                            entry = {'recetas':p['recetas'],'importe100':p['importe100'],'ac_instituto':p['ac_instituto']}
                            if 'MAMI' in plan_name: planes_data['MAMI'] = entry
                            elif 'MAYOR' in plan_name: planes_data['MAYOR COBERTURA'] = entry
                            elif 'AMPARO' in plan_name: planes_data['RECURSOS DE AMPARO'] = entry
                            elif 'DIRECTORIO' in plan_name: planes_data['RESOLUCIÓN DE DIRECTORIO'] = entry
                            elif 'VACUNA' in plan_name: planes_data['VACUNAS'] = entry
                    else:
                        pd_data = parse_json(ask_claude(
                            pdf_to_content(cb, 'CARÁTULA PLAN IOMA') + [{"type":"text","text":"Leer Convenio/Plan para identificar el plan. Extraé: {\"plan\":\"nombre del plan\",\"recetas\":0,\"importe100\":0.0,\"ac_instituto\":0.0}"}],
                            SYSTEM_JSON))
                        plan_name = pd_data.get('plan','').upper()
                        entry = {'recetas':pd_data['recetas'],'importe100':pd_data['importe100'],'ac_instituto':pd_data['ac_instituto']}
                        if 'MAMI' in plan_name: planes_data['MAMI'] = entry
                        elif 'MAYOR' in plan_name: planes_data['MAYOR COBERTURA'] = entry
                        elif 'AMPARO' in plan_name: planes_data['RECURSOS DE AMPARO'] = entry
                        elif 'DIRECTORIO' in plan_name: planes_data['RESOLUCIÓN DE DIRECTORIO'] = entry
                        elif 'VACUNA' in plan_name: planes_data['VACUNAS'] = entry

                opf_data = parse_json(ask_claude(
                    pdf_to_content(read_file(opf_file), 'OPF IOMA') + [{"type":"text","text":"Efvo.IOMA AMBULATORIO = anticipo. Sección RETENCIONES línea RGI = ing_brutos_anticipo. Extraé: {\"efvo_ioma\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":0,\"ing_brutos_anticipo\":0.0}"}],
                    SYSTEM_JSON))
                pre_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pre_file), 'PRE IOMA') + [{"type":"text","text":"Extraé: {\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"fdo_prest_colfarma\":0.0,\"nrf_ant\":0.0,\"nrf_def\":0.0,\"nrf_directas\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}"}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pago_file), 'PAGO FINAL IOMA') + [{"type":"text","text":"Sección RETENCIONES línea RGI = ing_brutos_pago. Extraé: {\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0,\"ing_brutos_pago\":0.0}"}],
                    SYSTEM_JSON))

                nr_data = {'nr_por_fecha':[]}
                if nr_file:
                    nr_text = xls_to_text(read_file(nr_file), os.path.basename(nr_file))
                    nr_data = parse_json(ask_claude(
                        [{"type":"text","text":f"NOTAS DE RECUPERO IOMA:\n{nr_text}\n\nSumá NRF y NRFD agrupados por fecha de Impresion. Extraé: {{\"nr_por_fecha\":[{{\"fecha\":\"DD/MM/YYYY\",\"monto\":0.0}}]}}"}],
                        SYSTEM_JSON))

                buf = build_ioma_excel(
                    {'planes':planes_data,'opf':opf_data,'pre':pre_data,'pago':pago_data,'nr':nr_data,'fecha_cierre':agudo_data['fecha_cierre']},
                    mes_num, anio[-2:])
                reportes.append((f"{anio[-2:]}.{mes_num} - Reporte IOMA.xlsx", buf.getvalue()))
            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_IOMA.zip"'})


@router.post("/batch/osde")
async def batch_osde(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []; errores = []
        meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}

        for root, dirs, files in os.walk(tmpdir):
            rel = os.path.relpath(root, tmpdir).replace('\\','/')
            parts = rel.split('/')
            if not any(m in rel for m in meses_map.keys()): continue
            if not any(os.path.isdir(os.path.join(root, d)) for d in os.listdir(root)): continue

            all_files = [os.path.join(root, f) for f in os.listdir(root) if os.path.isfile(os.path.join(root, f))]
            subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]
            liq_dir = next((d for d in subdirs if 'liquidac' in os.path.basename(d).lower()), None)
            if not liq_dir: continue

            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]
            nr_file = find_file(all_files, 'notas', 'recupero')
            car_file = find_file(liq_files, 'cierre', 'caratula', 'carátula')
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')

            if not all([car_file, pre_file, pago_file]):
                errores.append(f"❌ Faltan archivos en: {rel}\n   → Verificá que la carpeta tenga todos los archivos necesarios (Carátula, PRE, Pago, etc.)"); continue

            anio_part = next((p for p in parts if p.isdigit() and len(p)==4), None)
            mes_part = next((p for p in parts if any(m in p for m in meses_map.keys())), None)
            if not anio_part or not mes_part:
                errores.append(f"❌ No se pudo determinar mes/año en: {rel}\n   → Verificá que la estructura de carpetas sea: año/mes/quincena"); continue
            mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
            anio = anio_part

            try:
                car_data = parse_json(ask_claude(
                    pdf_to_content(read_file(car_file), 'CARÁTULA OSDE') + [{"type":"text","text":"Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"afiliado\":0.0,\"a_cargo_osde\":0.0,\"bonificacion\":0.0,\"total_verificar\":0.0}"}],
                    SYSTEM_JSON))
                pre_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pre_file), 'PRE OSDE') + [{"type":"text","text":"La tabla tiene columnas: Concepto, Base Cálculo, Créditos, Débitos. REGLAS: 1) Para retencion_fdo_res, ret_col_art12 y notas_credito tomá SIEMPRE el valor de la columna Débitos, NO la Base de Cálculo. 2) Para ajuste_facturacion: puede haber varias líneas de Ajuste Facturación. Identificá pares que se cancelan entre sí (mismo monto, uno en Débitos y otro en Créditos) y excluílos. El ajuste_facturacion es el monto de la línea que NO tiene contraparte que la cancele (si es débito es positivo, si es crédito es negativo). Si no hay ajuste real, ajuste_facturacion=0. 3) neto_cobrar = fila Neto a Cobrar columna Créditos. Extraé: {\"nro_liquidacion\":0,\"ajuste_facturacion\":0.0,\"retencion_fdo_res\":0.0,\"ret_col_art12\":0.0,\"notas_credito\":0.0,\"neto_cobrar\":0.0}"}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pago_file), 'PAGO FINAL OSDE') + [{"type":"text","text":f"La fecha_pago es la fecha del ENCABEZADO del documento (campo Fecha:), NO la fecha de presentación de la tabla. El nro_comprobante_pago es el número de Orden de Pago del encabezado. Confirmar que existe línea OSDE con liquidación nro {pre_data.get('nro_liquidacion','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}}"}],
                    SYSTEM_JSON))
                nr_data = {'nr_monto':0.0,'nr_fecha':'01/01/2000'}
                if nr_file:
                    nr_data = parse_json(ask_claude(
                        pdf_to_content(read_file(nr_file), 'NOTA DE CRÉDITO DEL SUD') + [{"type":"text","text":"Tomar el TOTAL del documento. Extraé: {\"nr_monto\":0.0,\"nr_fecha\":\"DD/MM/YYYY\"}"}],
                        SYSTEM_JSON))

                buf = build_osde_excel({'caratula':car_data,'pre':pre_data,'pago':pago_data,'nr':nr_data}, mes_num, anio[-2:])
                reportes.append((f"{anio[-2:]}.{mes_num} - Reporte OSDE.xlsx", buf.getvalue()))
            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_OSDE.zip"'})


@router.post("/batch/ospecon")
async def batch_ospecon(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []; errores = []
        meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}

        for root, dirs, files in os.walk(tmpdir):
            rel = os.path.relpath(root, tmpdir).replace('\\','/')
            parts = rel.split('/')
            if not any(m in rel for m in meses_map.keys()): continue
            if not any(os.path.isdir(os.path.join(root, d)) for d in os.listdir(root)): continue

            all_files = [os.path.join(root, f) for f in os.listdir(root) if os.path.isfile(os.path.join(root, f))]
            subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]
            liq_dir = next((d for d in subdirs if 'liquidac' in os.path.basename(d).lower()), None)
            if not liq_dir: continue

            car_files = [f for f in all_files if 'caratula' in os.path.basename(f).lower() or 'carátula' in os.path.basename(f).lower()]
            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')

            if not all([car_files, pre_file, pago_file]):
                errores.append(f"❌ Faltan archivos en: {rel}\n   → Verificá que la carpeta tenga todos los archivos necesarios (Carátula, PRE, Pago, etc.)"); continue

            anio_part = next((p for p in parts if p.isdigit() and len(p)==4), None)
            mes_part = next((p for p in parts if any(m in p for m in meses_map.keys())), None)
            if not anio_part or not mes_part:
                errores.append(f"❌ No se pudo determinar mes/año en: {rel}\n   → Verificá que la estructura de carpetas sea: año/mes/quincena"); continue
            mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
            anio = anio_part

            try:
                car_data = parse_json(ask_claude(
                    pdf_to_content(read_file(car_files[0]), 'CARÁTULA OSPECON') + [{"type":"text","text":"La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
                    SYSTEM_JSON))
                pre_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pre_file), 'PRE OSPECON') + [{"type":"text","text":"Columnas: Concepto, Base Cálculo, Créditos, Débitos. Para bonificacion, retencion_fdo_res y ret_col_art12 tomá SIEMPRE la columna Débitos. Para ajuste_facturacion: identificá pares que se cancelan entre sí y excluílos, el ajuste_facturacion es el monto de la línea sin contraparte (positivo si débito, negativo si crédito), 0 si no hay. neto_cobrar = fila Neto a Cobrar columna Créditos. Extraé: {\"nro_liquidacion\":0,\"ajuste_facturacion\":0.0,\"bonificacion\":0.0,\"retencion_fdo_res\":0.0,\"ret_col_art12\":0.0,\"neto_cobrar\":0.0}"}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pago_file), 'PAGO FINAL OSPECON') + [{"type":"text","text":f"La fecha_pago es la fecha del ENCABEZADO del documento (campo Fecha:). El nro_comprobante_pago es el número de Orden de Pago del encabezado. Confirmar línea OSPECON con liquidación nro {pre_data.get('nro_liquidacion','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":0}}"}],
                    SYSTEM_JSON))

                buf = build_ospecon_excel({'caratula':car_data,'pre':pre_data,'pago':pago_data}, mes_num, anio[-2:])
                reportes.append((f"{anio[-2:]}.{mes_num} - Reporte OSPECON.xlsx", buf.getvalue()))
            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_OSPECON.zip"'})


@router.post("/batch/osprera")
async def batch_osprera(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []; errores = []
        meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}

        for root, dirs, files in os.walk(tmpdir):
            rel = os.path.relpath(root, tmpdir).replace('\\','/')
            parts = rel.split('/')
            if not any(m in rel for m in meses_map.keys()): continue

            all_files = [os.path.join(root, f) for f in os.listdir(root) if os.path.isfile(os.path.join(root, f))]
            subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]

            anio_part = next((p for p in parts if p.isdigit() and len(p)==4), None)
            mes_part = next((p for p in parts if any(m in p for m in meses_map.keys())), None)
            if not anio_part or not mes_part: continue
            mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
            anio = anio_part

            # Detectar si tiene quincena
            con_quincena = any('1Q' in p or '2Q' in p for p in parts)
            quincena_val = next((p for p in parts if '1Q' in p or '2Q' in p), None)
            if quincena_val:
                quincena_val = '1Q' if '1Q' in quincena_val else '2Q'

            liq_dir = next((d for d in subdirs if 'liquidac' in os.path.basename(d).lower()), None)
            car_dir = next((d for d in subdirs if 'caratula' in os.path.basename(d).lower() or 'carátula' in os.path.basename(d).lower()), None)

            if con_quincena:
                # Archivos sueltos en la carpeta
                car_files = [f for f in all_files if any(x in os.path.basename(f).lower() for x in ['caratula','carátula'])]
                nr_file = find_file(all_files, 'notas', 'recupero', '.xls')
            else:
                car_files = sorted([os.path.join(car_dir, f) for f in os.listdir(car_dir)]) if car_dir else []
                nr_file = None

            if not liq_dir or not car_files: continue
            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')
            if not all([pre_file, pago_file]): continue

            try:
                planes_data = {'GENERAL':{'recetas':0,'importe100':0.0,'ac_os':0.0},'TRATAMIENTO PROLONGADO':{'recetas':0,'importe100':0.0,'ac_os':0.0},'MONOTRIBUTISTAS':{'recetas':0,'importe100':0.0,'ac_os':0.0},'RURAL':{'recetas':0,'importe100':0.0,'ac_os':0.0},'DECLARACIÓN DE DISPENSA':{'recetas':0,'importe100':0.0,'ac_os':0.0}}
                fecha_cierre = None
                for cp in car_files:
                    car_data = parse_json(ask_claude(
                        pdf_to_content(read_file(cp), 'CARÁTULA OSPRERA') + [{"type":"text","text":"Identificar el plan exacto leyendo el campo Convenio/Plan. La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"plan\":\"nombre completo del convenio/plan\",\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
                        SYSTEM_JSON))
                    if not fecha_cierre: fecha_cierre = car_data.get('fecha_cierre')
                    pn = car_data.get('plan','').upper()
                    k = 'MONOTRIBUTISTAS' if 'MONOT' in pn else 'RURAL' if 'RURAL' in pn else 'PROLONGADO' if 'PROLONGADO' in pn else 'DECLARACIÓN DE DISPENSA' if 'DISPENSA' in pn else 'GENERAL'
                    planes_data[k]['recetas'] += car_data.get('nro_recetas',0)
                    planes_data[k]['importe100'] += car_data.get('importe_total',0.0)
                    planes_data[k]['ac_os'] += car_data.get('ac_os',0.0)

                nr_data = None
                if con_quincena and nr_file:
                    nr_text = xls_to_text(read_file(nr_file), os.path.basename(nr_file))
                    nr_data = parse_json(ask_claude(
                        [{"type":"text","text":f"NOTAS DE RECUPERO OSPRERA:\n{nr_text}\n\nSumá NRF y NRFD agrupados por fecha. Extraé: {{\"nr_por_fecha\":[{{\"fecha\":\"DD/MM/YYYY\",\"monto\":0.0}}]}}"}],
                        SYSTEM_JSON))

                pre_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pre_file), 'PRE OSPRERA') + [{"type":"text","text":"Extraé: {\"fecha_presentacion\":\"DD/MM/YYYY\",\"nro_comprobante\":0,\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"notas_credito\":0.0,\"fdo_prest_colfarma\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. deb_cred_os = DEB/CRED DE OBRA SOCIAL (negativo si es débito). bonificaciones = BONIFICACIONES. notas_credito = NOTAS DE CREDITO. fdo_prest_colfarma = FDO PREST COLFARMA. total_liquidacion = Total liquidación."}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pago_file), 'PAGO FINAL OSPRERA') + [{"type":"text","text":f"La fecha_pago es la Fecha del encabezado del documento. El nro_comprobante_pago es el número de Comprobante del encabezado. Confirmar que existe línea OSPRERA con comprobante {pre_data.get('nro_comprobante','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":\"\"}}"}],
                    SYSTEM_JSON))

                q_str = f".{quincena_val}" if con_quincena and quincena_val else ""
                buf = build_osprera_excel(
                    {'planes':planes_data,'pre':pre_data,'pago':pago_data,'fecha_cierre':fecha_cierre,'opf':None,'nr':nr_data,'con_quincena':con_quincena,'quincena':quincena_val},
                    mes_num, anio[-2:])
                reportes.append((f"{anio[-2:]}.{mes_num}{q_str} - Reporte OSPRERA.xlsx", buf.getvalue()))
            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_OSPRERA.zip"'})


@router.post("/batch/unionpersonal")
async def batch_unionpersonal(zip_file: UploadFile = File(...)):
    zip_bytes = await zip_file.read()
    output_zip = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmpdir)

        reportes = []; errores = []
        meses_map = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}

        for root, dirs, files in os.walk(tmpdir):
            rel = os.path.relpath(root, tmpdir).replace('\\','/')
            parts = rel.split('/')
            if not any(m in rel for m in meses_map.keys()): continue
            if not any(os.path.isdir(os.path.join(root, d)) for d in os.listdir(root)): continue

            subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]
            car_dir = next((d for d in subdirs if 'caratula' in os.path.basename(d).lower() or 'carátula' in os.path.basename(d).lower()), None)
            liq_dir = next((d for d in subdirs if 'liquidac' in os.path.basename(d).lower()), None)
            if not car_dir or not liq_dir: continue

            car_files = sorted([os.path.join(car_dir, f) for f in os.listdir(car_dir) if f.endswith('.pdf')])
            liq_files = [os.path.join(liq_dir, f) for f in os.listdir(liq_dir)]
            opf_file = find_file(liq_files, 'opf')
            pre_file = find_file(liq_files, 'pre')
            pago_file = find_file(liq_files, 'pago')

            if not all([car_files, opf_file, pre_file, pago_file]):
                errores.append(f"❌ Faltan archivos en: {rel}\n   → Verificá que la carpeta tenga todos los archivos necesarios (Carátula, PRE, Pago, etc.)"); continue

            anio_part = next((p for p in parts if p.isdigit() and len(p)==4), None)
            mes_part = next((p for p in parts if any(m in p for m in meses_map.keys())), None)
            if not anio_part or not mes_part: continue
            mes_num = next((v for k,v in meses_map.items() if k in mes_part), None)
            anio = anio_part

            try:
                planes_data = {'PLANES VARIOS':{'recetas':0,'importe100':0.0,'ac_os':0.0},'DECLARACIÓN DE DISPENSA':{'recetas':0,'importe100':0.0,'ac_os':0.0}}
                fecha_cierre = None
                for cp in car_files:
                    car_data = parse_json(ask_claude(
                        pdf_to_content(read_file(cp), 'CARÁTULA UNIÓN PERSONAL') + [{"type":"text","text":"Identificar el plan: Planes Varios o Declaracion de dispensa. La fecha_cierre es la \"Fecha de generación\", NO la \"Fecha de Proceso\". Extraé: {\"plan\":\"nombre del plan\",\"fecha_cierre\":\"DD/MM/YYYY\",\"nro_recetas\":0,\"importe_total\":0.0,\"ac_os\":0.0}"}],
                        SYSTEM_JSON))
                    if not fecha_cierre: fecha_cierre = car_data.get('fecha_cierre')
                    key = 'DECLARACIÓN DE DISPENSA' if 'DISPENSA' in car_data.get('plan','').upper() else 'PLANES VARIOS'
                    planes_data[key]['recetas'] += car_data.get('nro_recetas',0)
                    planes_data[key]['importe100'] += car_data.get('importe_total',0.0)
                    planes_data[key]['ac_os'] += car_data.get('ac_os',0.0)

                opf_data = parse_json(ask_claude(
                    pdf_to_content(read_file(opf_file), 'OPF UNIÓN PERSONAL') + [{"type":"text","text":"Buscar línea UNION PERSONAL (SIFAR) con descripción que empiece con Efvo. La fecha_opf es la Fecha del encabezado. El nro_comprobante_opf es el Comprobante del encabezado. Extraé: {\"efvo_up\":0.0,\"fecha_opf\":\"DD/MM/YYYY\",\"nro_comprobante_opf\":\"\"}"}],
                    SYSTEM_JSON))
                pre_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pre_file), 'PRE UNIÓN PERSONAL') + [{"type":"text","text":"Extraé: {\"fecha_presentacion\":\"DD/MM/YYYY\",\"nro_comprobante\":0,\"deb_cred_os\":0.0,\"bonificaciones\":0.0,\"fdo_prest_colfarma\":0.0,\"retencion_colegio_art12\":0.0,\"total_liquidacion\":0.0}. deb_cred_os = DEB/CRED DE OBRA SOCIAL (negativo si es débito, 0 si no aparece). bonificaciones=BONIFICACIONES, fdo_prest_colfarma=FDO PREST COLFARMA, total_liquidacion=Total liquidación."}],
                    SYSTEM_JSON))
                pago_data = parse_json(ask_claude(
                    pdf_to_content(read_file(pago_file), 'PAGO FINAL UNIÓN PERSONAL') + [{"type":"text","text":f"La fecha_pago es la Fecha del encabezado. El nro_comprobante_pago es el Comprobante del encabezado. Buscar línea UNION PERSONAL con liquidación nro {pre_data.get('nro_comprobante','')}. Extraé: {{\"fecha_pago\":\"DD/MM/YYYY\",\"nro_comprobante_pago\":\"\"}}"}],
                    SYSTEM_JSON))

                buf = build_unionpersonal_excel(
                    {'planes':planes_data,'opf':opf_data,'pre':pre_data,'pago':pago_data,'fecha_cierre':fecha_cierre},
                    mes_num, anio[-2:])
                reportes.append((f"{anio[-2:]}.{mes_num} - Reporte Union Personal.xlsx", buf.getvalue()))
            except Exception as e:
                errores.append(f"❌ Error procesando: {rel}\n   Causa: {str(e)}\n   → Intentá generar este reporte de forma individual en la app")

        if not reportes:
            raise HTTPException(status_code=400, detail=f"No se generaron reportes. Errores: {errores}")
        with zipfile.ZipFile(output_zip, 'w') as zout:
            for filename, data in reportes:
                zout.writestr(filename, data)
            if errores:
                zout.writestr("errores.txt", "REPORTES CON ERRORES\n" + "="*50 + "\n\n" + "\n\n".join(errores) + "\n\n" + "="*50 + "\nPara los reportes con error, intentá generarlos de forma individual desde la app.")

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
                             headers={'Content-Disposition': 'attachment; filename="Reportes_UP.zip"'})

# ── REPORTE ANUAL ─────────────────────────────────────────────────────────────

def detectar_os_desde_nombre(filename):
    """Detecta la obra social y período desde el nombre del archivo"""
    name = filename.upper()
    if 'PAMI' in name: return 'PAMI'
    if 'IOMA' in name: return 'IOMA'
    if 'OSPECON' in name: return 'OSPECON'
    if 'OSPRERA' in name: return 'OSPRERA'
    if 'UNION' in name or 'PERSONAL' in name: return 'UNION PERSONAL'
    if 'OSDE' in name: return 'OSDE'
    # Formato antiguo sin nombre de OS: "25.01.1Q - Reporte.xlsx"
    if filename[:2].isdigit() and '.' in filename[:5]: return 'PAMI'
    return 'DESCONOCIDA'

def leer_resumen_reporte(xlsx_bytes, filename):
    """Lee la tab Resumen de un reporte generado y devuelve los datos"""
    import tempfile
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        if 'Resumen' not in wb.sheetnames:
            return None
        ws = wb['Resumen']
        # Leer todas las filas con datos
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)

        # Tabla 1: fila 1 = headers, fila 2 = datos (índices 0 y 1)
        # Tabla 2: fila 4 = headers, fila 5 = datos (índices 3 y 4)
        # Tabla 3: fila 7 = headers, fila 8 = datos (índices 6 y 7)
        def make_dict(header_row, data_row):
            if header_row is None or data_row is None:
                return {}
            return {str(k): v for k, v in zip(header_row, data_row) if k is not None}

        d1 = make_dict(rows[0] if len(rows) > 0 else None, rows[1] if len(rows) > 1 else None)
        d2 = make_dict(rows[3] if len(rows) > 3 else None, rows[4] if len(rows) > 4 else None)
        d3 = make_dict(rows[6] if len(rows) > 6 else None, rows[7] if len(rows) > 7 else None)

        # Crear un DataFrame dummy con la estructura esperada
        import pandas as pd
        combined_rows = [
            list(rows[0]) if len(rows) > 0 else [],
            list(rows[1]) if len(rows) > 1 else [],
            [None] * 15,
            list(rows[3]) if len(rows) > 3 else [],
            list(rows[4]) if len(rows) > 4 else [],
            [None] * 15,
            list(rows[6]) if len(rows) > 6 else [],
            list(rows[7]) if len(rows) > 7 else [],
        ]
        max_cols = max(len(r) for r in combined_rows) if combined_rows else 1
        padded = [r + [None]*(max_cols - len(r)) for r in combined_rows]
        df = pd.DataFrame(padded)

        return {'filename': filename, 'resumen_raw': df, 'd1': d1, 'd2': d2, 'd3': d3}
    finally:
        os.unlink(tmp_path)

def build_reporte_anual(reportes_data, os_nombre, anio):
    """Construye el reporte anual consolidado"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    DARK_BLUE_HEX = '1F3864'
    MID_BLUE_HEX = '2E5FA3'
    LIGHT_BLUE_HEX = 'D6E4F0'
    GREEN_HEX = '1E8449'
    WHITE_HEX = 'FFFFFF'
    ORANGE_HEX = 'E67E22'

    def header_cell(ws, cell, value, bg=DARK_BLUE_HEX, fg=WHITE_HEX, bold=True, size=10, halign='center'):
        c = ws[cell]
        c.value = value
        c.font = Font(bold=bold, color=fg, size=size, name='Arial')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)

    def data_cell(ws, cell, value, num_fmt=None, bold=False, bg=None, halign='right'):
        c = ws[cell]
        c.value = value
        c.font = Font(bold=bold, size=10, name='Arial')
        c.alignment = Alignment(horizontal=halign, vertical='center')
        if num_fmt: c.number_format = num_fmt
        if bg: c.fill = PatternFill('solid', fgColor=bg)

    def border_range(ws, min_row, max_row, min_col, max_col):
        thin = Side(style='thin', color='BFBFBF')
        medium = Side(style='medium', color=DARK_BLUE_HEX)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = Border(
                    top=medium if cell.row==min_row else thin,
                    bottom=medium if cell.row==max_row else thin,
                    left=medium if cell.column==min_col else thin,
                    right=medium if cell.column==max_col else thin
                )

    wb = Workbook()
    wb.remove(wb.active)

    # ── Procesar datos de los reportes ──────────────────────────────────────
    filas_principales = []
    filas_desglose = []
    filas_diferencias = []

    for rd in sorted(reportes_data, key=lambda x: x['filename']):
        periodo = rd['filename'].replace(' - Reporte.xlsx','').replace(' - Reporte PAMI.xlsx','').replace(' - Reporte IOMA.xlsx','').replace(' - Reporte OSDE.xlsx','').replace(' - Reporte OSPECON.xlsx','').replace(' - Reporte OSPRERA.xlsx','').replace(' - Reporte Union Personal.xlsx','').replace('.xlsx','')

        d1 = rd.get('d1', {})
        d2 = rd.get('d2', {})
        d3 = rd.get('d3', {})

        if d1:
            d1['PERIODO'] = periodo
            filas_principales.append(d1)
        if d2:
            d2['PERIODO'] = periodo
            filas_desglose.append(d2)
        if d3:
            d3['PERIODO'] = periodo
            filas_diferencias.append(d3)

    # ── TAB 1: Resumen Anual ────────────────────────────────────────────────
    ws1 = wb.create_sheet('Resumen Anual')
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = 'B3'

    # Título
    ws1.merge_cells('A1:P1')
    t = ws1['A1']
    t.value = f'REPORTE ANUAL {os_nombre} — {anio}'
    t.font = Font(bold=True, size=16, color=WHITE_HEX, name='Arial')
    t.fill = PatternFill('solid', fgColor=DARK_BLUE_HEX)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    cols1 = ['PERIODO','RECETAS','PVP','PVP PAMI','TOTAL','%PVP','AFILIADO','%AFL','PAMI','%PAMI','DIAS PROM.','DEBITOS','RETENCIONES','%RET','BONIFICACIONES','%BON']
    col_keys = ['PERIODO','RECETAS','PVP','PVP PAMI','TOTAL','%PVP','AFILIADO','%TOTAL AFL','PAMI','%TOTAL PAMI','DIAS PROM. PAGO','DEBITOS','RETENCIONES','%PVP PAMI RET','BONIFICACIONES','%PVP PAMI BON']

    for i, col in enumerate(cols1):
        header_cell(ws1, f'{get_column_letter(i+1)}2', col)
    ws1.row_dimensions[2].height = 30

    num_fmt_map = {
        'RECETAS':'#,##0', 'PVP':'#,##0', 'PVP PAMI':'#,##0', 'TOTAL':'#,##0',
        '%PVP':'0.0%', 'AFILIADO':'#,##0', '%AFL':'0.0%', 'PAMI':'#,##0',
        '%PAMI':'0.0%', 'DIAS PROM.':'#,##0.0', 'DEBITOS':'#,##0',
        'RETENCIONES':'#,##0', '%RET':'0.0%', 'BONIFICACIONES':'#,##0', '%BON':'0.0%'
    }

    for row_idx, d in enumerate(filas_principales):
        r = row_idx + 3
        alt_bg = 'EBF2FA' if row_idx % 2 == 0 else None
        for col_idx, (col_display, col_key) in enumerate(zip(cols1, col_keys)):
            cell_addr = f'{get_column_letter(col_idx+1)}{r}'
            val = d.get(col_key, '')
            fmt = num_fmt_map.get(col_display)
            halign = 'left' if col_display == 'PERIODO' else 'right'
            data_cell(ws1, cell_addr, val, num_fmt=fmt, bg=alt_bg, halign=halign)

    # Fila de totales/promedios
    if filas_principales:
        last_data_row = len(filas_principales) + 2
        total_row = last_data_row + 1
        ws1.row_dimensions[total_row].height = 20
        header_cell(ws1, f'A{total_row}', 'TOTAL / PROM.', bg=MID_BLUE_HEX)
        for col_idx, col_display in enumerate(cols1[1:], 2):
            col_letter = get_column_letter(col_idx)
            fmt = num_fmt_map.get(col_display)
            if col_display in ['%PVP','%AFL','%PAMI','%RET','%BON','DIAS PROM.']:
                formula = f'=AVERAGE({col_letter}3:{col_letter}{last_data_row})'
            else:
                formula = f'=SUM({col_letter}3:{col_letter}{last_data_row})'
            c = ws1[f'{col_letter}{total_row}']
            c.value = formula
            c.font = Font(bold=True, color=WHITE_HEX, size=10, name='Arial')
            c.fill = PatternFill('solid', fgColor=MID_BLUE_HEX)
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt: c.number_format = fmt

        border_range(ws1, 2, total_row, 1, len(cols1))

    # Anchos de columna
    ws1.column_dimensions['A'].width = 18
    for i in range(2, len(cols1)+1):
        ws1.column_dimensions[get_column_letter(i)].width = 14

    # ── TAB 2: Desglose de Pagos ────────────────────────────────────────────
    ws2 = wb.create_sheet('Desglose Pagos')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'B3'

    ws2.merge_cells('A1:M1')
    t2 = ws2['A1']
    t2.value = f'DESGLOSE DE PAGOS — {os_nombre} {anio}'
    t2.font = Font(bold=True, size=16, color=WHITE_HEX, name='Arial')
    t2.fill = PatternFill('solid', fgColor=DARK_BLUE_HEX)
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    cols2 = ['PERIODO','ANTICIPO','%ANT','DIAS ANT','LIQ FINAL','%LIQ','DIAS LIQ','NR','%NR','DIAS NR','EFVO DROG','%EFVO','DIAS EFVO']
    col_keys2 = ['PERIODO','ANTICIPO','%PAMI ANT','DIAS ANT','LIQ FINAL','%PAMI LIQ FINAL','DIAS LIQ FINAL','NOTAS DE RECUPERO','%PAMI NR','DIAS NR','EFVO DROG','%PAMI EFVO DROG','DIAS EFVO DROG']
    num_fmt2 = {'ANTICIPO':'#,##0','%ANT':'0.0%','DIAS ANT':'#,##0','LIQ FINAL':'#,##0','%LIQ':'0.0%','DIAS LIQ':'#,##0','NR':'#,##0','%NR':'0.0%','DIAS NR':'#,##0','EFVO DROG':'#,##0','%EFVO':'0.0%','DIAS EFVO':'#,##0'}

    for i, col in enumerate(cols2):
        header_cell(ws2, f'{get_column_letter(i+1)}2', col)
    ws2.row_dimensions[2].height = 30

    for row_idx, d in enumerate(filas_desglose):
        r = row_idx + 3
        alt_bg = 'EBF2FA' if row_idx % 2 == 0 else None
        for col_idx, (col_display, col_key) in enumerate(zip(cols2, col_keys2)):
            cell_addr = f'{get_column_letter(col_idx+1)}{r}'
            val = d.get(col_key, '')
            fmt = num_fmt2.get(col_display)
            halign = 'left' if col_display == 'PERIODO' else 'right'
            data_cell(ws2, cell_addr, val, num_fmt=fmt, bg=alt_bg, halign=halign)

    if filas_desglose:
        last_data_row2 = len(filas_desglose) + 2
        total_row2 = last_data_row2 + 1
        header_cell(ws2, f'A{total_row2}', 'TOTAL / PROM.', bg=MID_BLUE_HEX)
        for col_idx, col_display in enumerate(cols2[1:], 2):
            col_letter = get_column_letter(col_idx)
            fmt = num_fmt2.get(col_display)
            if col_display.startswith('%') or col_display.startswith('DIAS'):
                formula = f'=AVERAGE({col_letter}3:{col_letter}{last_data_row2})'
            else:
                formula = f'=SUM({col_letter}3:{col_letter}{last_data_row2})'
            c = ws2[f'{col_letter}{total_row2}']
            c.value = formula
            c.font = Font(bold=True, color=WHITE_HEX, size=10, name='Arial')
            c.fill = PatternFill('solid', fgColor=MID_BLUE_HEX)
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt: c.number_format = fmt
        border_range(ws2, 2, total_row2, 1, len(cols2))

    ws2.column_dimensions['A'].width = 18
    for i in range(2, len(cols2)+1):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    # ── TAB 3: Diferencias ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Diferencias')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'B3'

    ws3.merge_cells('A1:D1')
    t3 = ws3['A1']
    t3.value = f'DIFERENCIAS — {os_nombre} {anio}'
    t3.font = Font(bold=True, size=16, color=WHITE_HEX, name='Arial')
    t3.fill = PatternFill('solid', fgColor=DARK_BLUE_HEX)
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 30

    # Detectar columnas de diferencias disponibles
    dif_cols_available = list(filas_diferencias[0].keys()) if filas_diferencias else []
    dif_cols_display = [c for c in dif_cols_available if c != 'PERIODO']

    headers3 = ['PERIODO'] + dif_cols_display
    for i, col in enumerate(headers3):
        header_cell(ws3, f'{get_column_letter(i+1)}2', col)
    ws3.row_dimensions[2].height = 30

    for row_idx, d in enumerate(filas_diferencias):
        r = row_idx + 3
        alt_bg = 'EBF2FA' if row_idx % 2 == 0 else None
        for col_idx, col in enumerate(headers3):
            cell_addr = f'{get_column_letter(col_idx+1)}{r}'
            val = d.get(col, '')
            halign = 'left' if col == 'PERIODO' else 'right'
            c = ws3[cell_addr]
            c.value = val
            c.font = Font(size=10, name='Arial')
            c.alignment = Alignment(horizontal=halign, vertical='center')
            c.number_format = '#,##0.00'
            if alt_bg and col != 'PERIODO': c.fill = PatternFill('solid', fgColor=alt_bg)
            # Colorear diferencias: rojo si negativo, verde si positivo
            if col != 'PERIODO' and isinstance(val, (int, float)):
                if val < -100: c.font = Font(size=10, name='Arial', color='C0392B', bold=True)
                elif val > 100: c.font = Font(size=10, name='Arial', color='1E8449', bold=True)

    if filas_diferencias:
        last_data_row3 = len(filas_diferencias) + 2
        total_row3 = last_data_row3 + 1
        header_cell(ws3, f'A{total_row3}', 'TOTAL', bg=MID_BLUE_HEX)
        for col_idx, col in enumerate(headers3[1:], 2):
            col_letter = get_column_letter(col_idx)
            c = ws3[f'{col_letter}{total_row3}']
            c.value = f'=SUM({col_letter}3:{col_letter}{last_data_row3})'
            c.font = Font(bold=True, color=WHITE_HEX, size=10, name='Arial')
            c.fill = PatternFill('solid', fgColor=MID_BLUE_HEX)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.number_format = '#,##0.00'
        border_range(ws3, 2, total_row3, 1, len(headers3))

    ws3.column_dimensions['A'].width = 18
    for i in range(2, len(headers3)+1):
        ws3.column_dimensions[get_column_letter(i)].width = 18

    # ── TAB 4: Gráficos ────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Gráficos')
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:N1')
    t4 = ws4['A1']
    t4.value = f'GRÁFICOS — {os_nombre} {anio}'
    t4.font = Font(bold=True, size=16, color=WHITE_HEX, name='Arial')
    t4.fill = PatternFill('solid', fgColor=DARK_BLUE_HEX)
    t4.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 30

    n = len(filas_principales)
    if n > 0:
        # Datos auxiliares para gráficos
        ws4['A3'] = 'PERIODO'; ws4['B3'] = 'PVP PAMI'; ws4['C3'] = 'TOTAL PAGADO'; ws4['D3'] = 'DIAS PROM'
        ws4['E3'] = 'ANTICIPO'; ws4['F3'] = 'LIQ FINAL'; ws4['G3'] = 'NR'; ws4['H3'] = 'EFVO DROG'
        for cell in [ws4['A3'],ws4['B3'],ws4['C3'],ws4['D3'],ws4['E3'],ws4['F3'],ws4['G3'],ws4['H3']]:
            cell.font = Font(bold=True, size=9, color=WHITE_HEX, name='Arial')
            cell.fill = PatternFill('solid', fgColor=MID_BLUE_HEX)
            cell.alignment = Alignment(horizontal='center')

        for i, (d1, d2) in enumerate(zip(filas_principales, filas_desglose)):
            r = i + 4
            ws4[f'A{r}'] = d1.get('PERIODO','')
            ws4[f'B{r}'] = d1.get('PVP PAMI', 0)
            ws4[f'C{r}'] = d1.get('PAMI', 0)
            ws4[f'D{r}'] = d1.get('DIAS PROM. PAGO', 0)
            ws4[f'E{r}'] = d2.get('ANTICIPO', 0)
            ws4[f'F{r}'] = d2.get('LIQ FINAL', 0)
            ws4[f'G{r}'] = d2.get('NOTAS DE RECUPERO', 0)
            ws4[f'H{r}'] = d2.get('EFVO DROG', 0)
            for col in 'BCDEFGH':
                ws4[f'{col}{r}'].number_format = '#,##0'
            ws4[f'D{r}'].number_format = '#,##0.0'

        # Gráfico 1: PVP PAMI vs Total Pagado (líneas)
        chart1 = LineChart()
        chart1.title = 'PVP PAMI vs Total Pagado'
        chart1.style = 10
        chart1.y_axis.title = 'Monto ($)'
        chart1.x_axis.title = 'Período'
        chart1.height = 12; chart1.width = 20

        data_pvp = Reference(ws4, min_col=2, max_col=3, min_row=3, max_row=3+n)
        cats = Reference(ws4, min_col=1, min_row=4, max_row=3+n)
        chart1.add_data(data_pvp, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.series[0].graphicalProperties.line.solidFill = MID_BLUE_HEX
        chart1.series[1].graphicalProperties.line.solidFill = GREEN_HEX
        ws4.add_chart(chart1, 'A6')

        # Gráfico 2: Composición de pagos (barras apiladas)
        chart2 = BarChart()
        chart2.type = 'col'; chart2.grouping = 'stacked'; chart2.overlap = 100
        chart2.title = 'Composición de Pagos'
        chart2.y_axis.title = 'Monto ($)'
        chart2.x_axis.title = 'Período'
        chart2.height = 12; chart2.width = 20

        data_comp = Reference(ws4, min_col=5, max_col=8, min_row=3, max_row=3+n)
        chart2.add_data(data_comp, titles_from_data=True)
        chart2.set_categories(cats)
        colors = [MID_BLUE_HEX, '5B9BD5', GREEN_HEX, ORANGE_HEX]
        for i, color in enumerate(colors):
            if i < len(chart2.series):
                chart2.series[i].graphicalProperties.solidFill = color
        ws4.add_chart(chart2, 'K6')

        # Gráfico 3: Días promedio de pago (barras)
        chart3 = BarChart()
        chart3.type = 'col'
        chart3.title = 'Días Promedio de Pago'
        chart3.y_axis.title = 'Días'
        chart3.x_axis.title = 'Período'
        chart3.height = 12; chart3.width = 20

        data_dias = Reference(ws4, min_col=4, max_col=4, min_row=3, max_row=3+n)
        chart3.add_data(data_dias, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.series[0].graphicalProperties.solidFill = MID_BLUE_HEX
        ws4.add_chart(chart3, 'A28')

        # Gráfico 4: Diferencias (barras)
        if filas_diferencias:
            ws4[f'A{4+n+2}'] = 'PERIODO'
            dif_keys = [c for c in filas_diferencias[0].keys() if c != 'PERIODO']
            for j, k in enumerate(dif_keys):
                ws4[f'{get_column_letter(j+2)}{4+n+2}'] = k
                for i, d in enumerate(filas_diferencias):
                    ws4[f'{get_column_letter(j+2)}{4+n+3+i}'] = d.get(k, 0)
            for i, d in enumerate(filas_diferencias):
                ws4[f'A{4+n+3+i}'] = d.get('PERIODO','')

            chart4 = BarChart()
            chart4.type = 'col'
            chart4.title = 'Diferencias'
            chart4.y_axis.title = 'Monto ($)'
            chart4.height = 12; chart4.width = 20
            dif_start_row = 4+n+2
            data_dif = Reference(ws4, min_col=2, max_col=1+len(dif_keys), min_row=dif_start_row, max_row=dif_start_row+len(filas_diferencias))
            cats_dif = Reference(ws4, min_col=1, min_row=dif_start_row+1, max_row=dif_start_row+len(filas_diferencias))
            chart4.add_data(data_dif, titles_from_data=True)
            chart4.set_categories(cats_dif)
            ws4.add_chart(chart4, 'K28')

    for col in 'ABCDEFGH':
        ws4.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf



@router.post("/reporte-anual")
async def reporte_anual(archivos: list[UploadFile] = File(...)):
    if not archivos:
        raise HTTPException(status_code=400, detail="No se subieron archivos")

    reportes_por_os = {}

    for archivo in archivos:
        xlsx_bytes = await archivo.read()
        os_nombre = detectar_os_desde_nombre(archivo.filename)
        resultado = leer_resumen_reporte(xlsx_bytes, archivo.filename)
        if resultado:
            if os_nombre not in reportes_por_os:
                reportes_por_os[os_nombre] = []
            reportes_por_os[os_nombre].append(resultado)

    if not reportes_por_os:
        raise HTTPException(status_code=400, detail="No se pudieron leer los archivos")

    # Detectar año desde los nombres de archivo
    import re
    primer_archivo = archivos[0].filename
    anio_match = re.search(r'20(\d{2})', primer_archivo)
    anio = f"20{anio_match.group(1)}" if anio_match else "2025"

    # Si hay una sola OS, devolver un Excel directo
    if len(reportes_por_os) == 1:
        os_nombre = list(reportes_por_os.keys())[0]
        reportes = reportes_por_os[os_nombre]
        buf = build_reporte_anual(reportes, os_nombre, anio)
        filename = f"Reporte Anual {os_nombre} {anio}.xlsx"
        return StreamingResponse(buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'})

    # Si hay múltiples OS, devolver un ZIP
    output_zip = io.BytesIO()
    with zipfile.ZipFile(output_zip, 'w') as zout:
        for os_nombre, reportes in reportes_por_os.items():
            buf = build_reporte_anual(reportes, os_nombre, anio)
            zout.writestr(f"Reporte Anual {os_nombre} {anio}.xlsx", buf.getvalue())

    output_zip.seek(0)
    return StreamingResponse(output_zip, media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename="Reportes Anuales {anio}.zip"'})
