from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
import anthropic
import io
import os
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import base64
import json
import httpx

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

DARK_BLUE='1F3864'; MID_BLUE='2E5FA3'; LIGHT_BLUE='D6E4F0'; WHITE='FFFFFF'; GREEN='1E8449'; ORANGE='C0392B'

def parse_date(s):
    if not s: return datetime.now()
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
        try: return datetime.strptime(s.strip(), fmt)
        except: pass
    return datetime.now()

def days_diff(d1, d2):
    return (d2 - d1).days

def ask_claude(content, system):
    msg = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        system=system,
        messages=[{"role": "user", "content": content}]
    )
    return msg.content[0].text

def pdf_to_content(pdf_bytes, label):
    b64 = base64.standard_b64encode(pdf_bytes).decode()
    return [
        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
        {"type": "text", "text": f"Documento: {label}"}
    ]

def parse_json(text):
    text = text.strip()
    start = text.find('{')
    end = text.rfind('}') + 1
    if start < 0 or end <= start:
        raise ValueError(f"No JSON found in: {text[:200]}")
    json_str = text[start:end]
    # Intentar parsear directo
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        pass
    # Limpiar caracteres problemáticos y reintentar
    import re
    # Eliminar saltos de línea dentro de strings
    json_str = re.sub(r'(?<!\)\n', ' ', json_str)
    json_str = re.sub(r'(?<!\)\r', ' ', json_str)
    # Eliminar caracteres de control
    json_str = re.sub(r'[-]', ' ', json_str)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON parse error: {e} in: {json_str[:200]}")

def xls_to_text(file_bytes, filename):
    import tempfile
    # Detectar si es XML/HTML disfrazado de XLS (común en sistemas de obras sociales)
    sample = file_bytes[:20]
    is_xml = (sample.startswith(b'\xff\xfe') or sample.startswith(b'<?xml') or
              sample.startswith(b'<') or b'<html' in sample[:100].lower() or
              b'<HTML' in sample[:100])
    if is_xml:
        # Decodificar como XML/HTML y extraer texto
        try:
            text = file_bytes.decode('utf-16')
        except Exception:
            try:
                text = file_bytes.decode('utf-8', errors='replace')
            except Exception:
                text = file_bytes.decode('latin-1', errors='replace')
        # Usar pandas para leer HTML
        import io as _io
        try:
            tables = pd.read_html(_io.StringIO(text))
            if tables:
                return '\n'.join(df.to_csv(index=False) for df in tables)
        except Exception:
            pass
        # Fallback: devolver texto plano sin tags
        import re
        clean = re.sub(r'<[^>]+>', ' ', text)
        clean = re.sub(r'\s+', ' ', clean)
        return clean
    suffix = '.xlsx' if filename.endswith('.xlsx') else '.xls'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        engine = "xlrd" if suffix == ".xls" else "openpyxl"
        df = pd.read_excel(tmp_path, header=None, engine=engine)
        return df.to_csv(index=False)
    finally:
        os.unlink(tmp_path)

SYSTEM_JSON = 'CRÍTICO: Respondé ÚNICAMENTE con el objeto JSON solicitado. Sin texto, sin explicaciones, sin markdown. Empezá con { y terminá con }.'

# ── EXCEL BUILDERS ────────────────────────────────────────────────────────────

def c(ws, coord, val, bold=False, size=10, color='000000', fill=None, halign='left', num_fmt=None):
    cell = ws[coord]
    cell.value = val
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=halign, vertical='center')
    if fill: cell.fill = PatternFill('solid', fgColor=fill)
    if num_fmt: cell.number_format = num_fmt

def n(ws, coord, val):
    c(ws, coord, val, halign='right', num_fmt='#,##0.00')

def ni(ws, coord, val):
    """Número entero sin decimales"""
    c(ws, coord, round(val) if val else 0, halign='right', num_fmt='#,##0')

def d(ws, coord, val):
    ws[coord].value = val
    ws[coord].number_format = 'DD/MM/YYYY'
    ws[coord].font = Font(size=10)
    ws[coord].alignment = Alignment(horizontal='right', vertical='center')

def box(ws, min_r, min_c, max_r, max_c):
    s = Side(style='medium', color=DARK_BLUE)
    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for cell in row:
            t = s if cell.row == min_r else None
            b = s if cell.row == max_r else None
            l = s if cell.column == min_c else None
            r = s if cell.column == max_c else None
            cell.border = Border(top=t, bottom=b, left=l, right=r)

def setup_ws(ws):
    ws.sheet_view.showGridLines = False
    for col, w in {'A':2,'B':20,'C':16,'D':3,'E':22,'F':16,'G':3,'H':22,'I':19,'J':3,'K':16,'L':14}.items():
        ws.column_dimensions[col].width = w
    for r, h in {1:12,2:18,3:26,4:12,5:18,6:26,7:12,8:18,9:26,10:12}.items():
        ws.row_dimensions[r].height = h
    for r in range(11, 50):
        ws.row_dimensions[r].height = 16

def header_bg(ws):
    for r in range(2, 4):
        for col in range(2, 13):
            ws.cell(r, col).fill = PatternFill('solid', fgColor=DARK_BLUE)
    for col in [4, 10, 12]:
        for r in [2, 3]:
            ws.cell(r, col).fill = PatternFill('solid', fgColor=DARK_BLUE)

def add_resumen_table(ws2, headers, data, table_name, ref, style):
    for i, h in enumerate(headers):
        ws2.cell(1, i+1, h)
    for i, v in enumerate(data):
        cell = ws2.cell(2, i+1)
        cell.value = v
        if isinstance(v, float) and abs(v) < 10:
            cell.number_format = '0.00%'
        else:
            cell.number_format = '#,##0.00'
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = style
    ws2.add_table(tbl)

# ── PAMI ──────────────────────────────────────────────────────────────────────

def build_pami_excel(data, q, mes, anio):
    car = data['caratula']; opf = data['opf']; pre = data['pre']; pago = data['pago']; nr = data['nr']

    fecha_pres = parse_date(car['fecha_cierre'])
    dias_ant = days_diff(fecha_pres, parse_date(opf['fecha_opf']))
    dias_liq = days_diff(fecha_pres, parse_date(pago['fecha_pago']))
    dias_nr = days_diff(fecha_pres, parse_date(nr.get('fecha_nr', car['fecha_cierre'])))
    dias_efsa = days_diff(fecha_pres, parse_date(nr.get('fecha_efsa', car['fecha_cierre'])))

    CCF=nr.get('nr_ccf',0); CCFD=nr.get('nr_ccfd',0); NAF=nr.get('nr_naf',0)
    NRFD=nr.get('nr_nrfd',0); EfSa=nr.get('nr_efsa',0)
    total_ccf_ccfd = CCF+CCFD; total_naf_nrfd = NAF+NRFD
    liq_final = pre['total_liquidacion'] - opf['efvo_pami']
    total_pagado = opf['efvo_pami'] + liq_final + total_ccf_ccfd + total_naf_nrfd + EfSa
    afiliado = car['total_pvp_pami'] - car['importe_bruto_convenio']
    total = afiliado + total_pagado
    deb_os = abs(pre['deb_cred_os']) if pre['deb_cred_os'] < 0 else 0
    cred_os = pre['deb_cred_os'] if pre['deb_cred_os'] > 0 else 0
    bonificaciones = abs(pre['bonif_ambulatorio']) + abs(pre['bonif_tiras']) + abs(pre['bonif_insulinas'])
    retenciones = abs(pre['ret_gtos_adm_cofa']) + abs(pre['retencion_colegio_art12']) + abs(pre['fdo_prest_colfarma'])
    notas_cred = abs(pre['nota_cred_ambulatorio']) + abs(pre['nota_cred_insulina']) + abs(pre['nota_cred_tiras'])
    diferencia_total = car['total_pvp'] - car['total_pvp_pami']
    pct70 = diferencia_total * 0.7; pct30 = diferencia_total * 0.3
    dif_nr = total_naf_nrfd - notas_cred
    dif_efvo = EfSa - abs(pre['efectivo_drogueria'])
    dif_ccf = total_ccf_ccfd - pct70
    dias_prom = (opf['efvo_pami']*dias_ant + liq_final*dias_liq + (total_ccf_ccfd+total_naf_nrfd)*dias_nr + EfSa*dias_efsa) / total_pagado
    periodo = f'{q}/{mes}/{anio}'

    wb = Workbook()
    ws = wb.active; ws.title = 'Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','PAMI',bold=True,size=28,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',car['nro_recetas'],bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_prom,1),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0.0')

    ws.merge_cells('B5:C5'); c(ws,'B5','TOTAL PVP',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('B6:C6'); ni(ws,'B6',car['total_pvp']); ws['B6'].font=Font(bold=True,size=13,color=WHITE); ws['B6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['B6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('E5:F5'); c(ws,'E5','PVP PAMI',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('E6:F6'); ni(ws,'E6',car['total_pvp_pami']); ws['E6'].font=Font(bold=True,size=13,color=WHITE); ws['E6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['E6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('H5:I5'); c(ws,'H5','AFILIADO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('H6:I6'); ni(ws,'H6',afiliado); ws['H6'].font=Font(bold=True,size=13,color=WHITE); ws['H6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['H6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('K5:L5'); c(ws,'K5','TOTAL PAGADO PAMI',size=9,color=WHITE,fill=GREEN,halign='center')
    ws.merge_cells('K6:L6'); ni(ws,'K6',total_pagado); ws['K6'].font=Font(bold=True,size=13,color=WHITE); ws['K6'].fill=PatternFill('solid',fgColor=GREEN); ws['K6'].alignment=Alignment(horizontal='center',vertical='center')

    for col,col_dias,lbl,dias,val in [('B','C','ANTICIPO',dias_ant,opf['efvo_pami']),('E','F','LIQ. FINAL',dias_liq,liq_final),('H','I','NOTAS RECUP.',dias_nr,total_ccf_ccfd+total_naf_nrfd),('K','L','EFVO. DROG.',dias_efsa,EfSa)]:
        c(ws,f'{col}8',lbl,size=9,color=WHITE,fill=MID_BLUE,halign='center')
        c(ws,f'{col_dias}8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        n(ws,f'{col}9',val); ws[f'{col}9'].font=Font(bold=True,size=13,color=WHITE); ws[f'{col}9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws[f'{col}9'].alignment=Alignment(horizontal='center',vertical='center')
        c(ws,f'{col_dias}9',dias,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','PAGOS A FARMACIA',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')
    ws.merge_cells('H11:L11'); c(ws,'H11','PAGOS A DROGUERIA',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','ANTICIPO (OPF)',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','TOTAL',bold=True); n(ws,'C13',opf['efvo_pami'])
    c(ws,'B14','Fecha pago'); d(ws,'C14',parse_date(opf['fecha_opf']))
    c(ws,'B15','Comprobante'); ws['C15'].value=opf['nro_comprobante_opf']; ws['C15'].alignment=Alignment(horizontal='right',vertical='center'); ws['C15'].font=Font(size=10)
    ws.merge_cells('B16:C16'); c(ws,'B16','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B17','Bruto a pagar antes imp.:'); n(ws,'C17',pre['total_liquidacion'])
    c(ws,'B18','TOTAL',bold=True); n(ws,'C18',liq_final)
    c(ws,'B19','Fecha pago'); d(ws,'C19',parse_date(pago['fecha_pago']))
    c(ws,'B20','Comprobante'); ws['C20'].value=pago['nro_comprobante_pago']; ws['C20'].alignment=Alignment(horizontal='right',vertical='center'); ws['C20'].font=Font(size=10)
    box(ws,11,2,20,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Ambulatorio:'); n(ws,'F13',abs(pre['bonif_ambulatorio']))
    c(ws,'E14','Tiras:'); n(ws,'F14',abs(pre['bonif_tiras']))
    c(ws,'E15','Insulinas:'); n(ws,'F15',abs(pre['bonif_insulinas']))
    c(ws,'E16','TOTAL',bold=True); n(ws,'F16',bonificaciones)
    ws.merge_cells('E17:F17'); c(ws,'E17','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E18','Gastos Adm. COFA:'); n(ws,'F18',abs(pre['ret_gtos_adm_cofa']))
    c(ws,'E19','Colegio Art. 12 SU:'); n(ws,'F19',abs(pre['retencion_colegio_art12']))
    c(ws,'E20','Fdo Prest. COLFARMA:'); n(ws,'F20',abs(pre['fdo_prest_colfarma']))
    c(ws,'E21','TOTAL',bold=True); n(ws,'F21',retenciones)
    ws.merge_cells('E22:F22'); c(ws,'E22','DÉB. / CRÉD. OS',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E23','Débito OS:'); n(ws,'F23',deb_os)
    c(ws,'E24','Crédito OS:'); n(ws,'F24',cred_os)
    box(ws,11,5,24,6)

    ws.merge_cells('H12:L12'); c(ws,'H12','NOTAS DE CRÉDITO',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'I13','Monto Calculado',bold=True,halign='center'); c(ws,'K13','Monto según NR',bold=True,halign='center'); c(ws,'L13','Diferencia',bold=True,halign='center')
    c(ws,'H14','Ambulatorio:'); n(ws,'I14',abs(pre['nota_cred_ambulatorio']))
    c(ws,'H15','Tiras:'); n(ws,'I15',abs(pre['nota_cred_tiras']))
    c(ws,'H16','Insulinas:'); n(ws,'I16',abs(pre['nota_cred_insulina']))
    c(ws,'H17','TOTAL',bold=True); n(ws,'I17',notas_cred); n(ws,'K17',total_naf_nrfd); n(ws,'L17',dif_nr)
    ws.merge_cells('H18:L18'); c(ws,'H18','70% / 30%',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'I19','Monto Calculado',bold=True,halign='center'); c(ws,'K19','Monto según NR',bold=True,halign='center'); c(ws,'L19','Diferencia',bold=True,halign='center')
    c(ws,'H20','70% CCF/CCFD',bold=True); n(ws,'I20',pct70); n(ws,'K20',total_ccf_ccfd); n(ws,'L20',dif_ccf)
    c(ws,'H21','30% Pérdida'); n(ws,'I21',pct30)
    ws.merge_cells('H22:L22'); c(ws,'H22','EFECTIVO DROGUERÍA',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'I23','Monto Calculado',bold=True,halign='center'); c(ws,'K23','Monto según NR',bold=True,halign='center'); c(ws,'L23','Diferencia',bold=True,halign='center')
    c(ws,'H24','TOTAL',bold=True); n(ws,'I24',abs(pre['efectivo_drogueria'])); n(ws,'K24',EfSa); n(ws,'L24',dif_efvo)
    box(ws,11,8,24,12)

    ws2 = wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines = False
    style = TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)

    h1=['RECETAS','PVP','PVP PAMI','TOTAL','%PVP','AFILIADO','%TOTAL AFL','PAMI','%TOTAL PAMI','DIAS PROM. PAGO','DEBITOS','RETENCIONES','%PVP PAMI RET','BONIFICACIONES','%PVP PAMI BON']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[car['nro_recetas'],car['total_pvp'],car['total_pvp_pami'],total,total/car['total_pvp'],afiliado,afiliado/total,total_pagado,total_pagado/total,round(dias_prom,2),deb_os,retenciones,retenciones/car['total_pvp_pami'],bonificaciones,bonificaciones/car['total_pvp_pami']]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_reporte',ref='A1:O2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)

    h2=['ANTICIPO','%PAMI ANT','DIAS ANT','LIQ FINAL','%PAMI LIQ FINAL','DIAS LIQ FINAL','NOTAS DE RECUPERO','%PAMI NR','DIAS NR','EFVO DROG','%PAMI EFVO DROG','DIAS EFVO DROG']
    for i,h in enumerate(h2): ws2.cell(4,i+1,h)
    row5=[opf['efvo_pami'],opf['efvo_pami']/total_pagado,dias_ant,liq_final,liq_final/total_pagado,dias_liq,total_ccf_ccfd+total_naf_nrfd,(total_ccf_ccfd+total_naf_nrfd)/total_pagado,dias_nr,EfSa,EfSa/total_pagado,dias_efsa]
    for i,v in enumerate(row5):
        cell=ws2.cell(5,i+1); cell.value=v
        cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl2=Table(displayName='tbl_desglose',ref='A4:L5'); tbl2.tableStyleInfo=style; ws2.add_table(tbl2)

    for i,h in enumerate(['Diferencias NR','Diferencias EFVO DROG','Diferencias CCF']): ws2.cell(7,i+1,h)
    for i,v in enumerate([dif_nr,dif_efvo,dif_ccf]):
        cell=ws2.cell(8,i+1); cell.value=v; cell.number_format='#,##0.00'
    tbl3=Table(displayName='tbl_diferencias',ref='A7:C8'); tbl3.tableStyleInfo=style; ws2.add_table(tbl3)
    for col in 'ABCDEFGHIJKLMNO': ws2.column_dimensions[col].width=20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── IOMA ──────────────────────────────────────────────────────────────────────

def build_ioma_excel(data, mes, anio):
    planes = data['planes']; opf = data['opf']; pre = data['pre']; pago = data['pago']; nr_data = data['nr']

    total_recetas = sum(p.get('recetas',0) for p in planes.values())
    total_importe100 = sum(p.get('importe100',0) for p in planes.values())
    total_ac = sum(p.get('ac_instituto',0) for p in planes.values())
    afiliado = total_importe100 - total_ac
    fecha_pres = parse_date(data['fecha_cierre'])

    total_nr=0; total_pond=0
    nr_por_fecha = nr_data.get('nr_por_fecha',[])
    for item in nr_por_fecha:
        m=item.get('monto',0); total_nr+=m
        total_pond+=m*days_diff(fecha_pres,parse_date(item.get('fecha','')))
    dias_nr_pond = total_pond/total_nr if total_nr else 0

    total_ing_brutos = opf['ing_brutos_anticipo'] + pago['ing_brutos_pago']
    liq_final = pre['total_liquidacion'] - opf['efvo_ioma'] - total_ing_brutos
    total_pagado = opf['efvo_ioma'] + liq_final + total_nr
    deb_os = abs(pre['deb_cred_os']) if pre['deb_cred_os']<0 else 0
    cred_os = pre['deb_cred_os'] if pre['deb_cred_os']>0 else 0
    ret_cofa = abs(pre['retencion_colegio_art12']) + abs(pre['fdo_prest_colfarma'])
    notas_cred = abs(pre['nrf_ant']) + abs(pre['nrf_def']) + abs(pre['nrf_directas'])
    dif_nr = total_nr - notas_cred
    dias_ant = days_diff(fecha_pres, parse_date(opf['fecha_opf']))
    dias_liq = days_diff(fecha_pres, parse_date(pago['fecha_pago']))
    dias_prom = (opf['efvo_ioma']*dias_ant + liq_final*dias_liq + total_nr*dias_nr_pond) / total_pagado
    periodo = f'{mes}/{anio}'
    planes_activos = {k:v for k,v in planes.items() if v.get('recetas',0)>0}

    wb = Workbook()
    ws = wb.active; ws.title = 'Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','IOMA',bold=True,size=28,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',total_recetas,bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_prom,1),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0.0')

    for coord,label,val,clr in [('B5:C5','IMPORTE 100%',total_importe100,MID_BLUE),('E5:F5','A/C INSTITUTO',total_ac,MID_BLUE),('H5:I5','AFILIADO',afiliado,MID_BLUE),('K5:L5','TOTAL PAGADO IOMA',total_pagado,GREEN)]:
        ws.merge_cells(coord); start=coord.split(':')[0]; end=coord.split(':')[1]
        r=int(start[1]); c_letter=start[0]; end_letter=end[0]
        c(ws,f'{c_letter}{r}',label,size=9,color=WHITE,fill=clr,halign='center')
        coord6=f'{c_letter}{r+1}:{end_letter}{r+1}'; ws.merge_cells(coord6)
        ni(ws,f'{c_letter}{r+1}',val); ws[f'{c_letter}{r+1}'].font=Font(bold=True,size=13,color=WHITE); ws[f'{c_letter}{r+1}'].fill=PatternFill('solid',fgColor=clr); ws[f'{c_letter}{r+1}'].alignment=Alignment(horizontal='center',vertical='center')

    for col,col_dias,lbl,dias,val in [('B','C','ANTICIPO',dias_ant,opf['efvo_ioma']),('E','F','LIQ. FINAL',dias_liq,liq_final),('H','I','NOTAS RECUP.',round(dias_nr_pond,1),total_nr)]:
        c(ws,f'{col}8',lbl,size=9,color=WHITE,fill=MID_BLUE,halign='center')
        c(ws,f'{col_dias}8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        n(ws,f'{col}9',val); ws[f'{col}9'].font=Font(bold=True,size=13,color=WHITE); ws[f'{col}9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws[f'{col}9'].alignment=Alignment(horizontal='center',vertical='center')
        c(ws,f'{col_dias}9',dias,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('K8:L8'); c(ws,'K8','ING. BRUTOS RETENIDOS',size=9,color=WHITE,fill=ORANGE,halign='center')
    ws.merge_cells('K9:L9'); ni(ws,'K9',total_ing_brutos); ws['K9'].font=Font(bold=True,size=13,color=WHITE); ws['K9'].fill=PatternFill('solid',fgColor=ORANGE); ws['K9'].alignment=Alignment(horizontal='center',vertical='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','CARÁTULAS',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')
    ws.merge_cells('H11:L11'); c(ws,'H11','PAGOS A DROGUERIA',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','COMPOSICIÓN POR PLAN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','PLAN',bold=True); c(ws,'C13','RECETAS',bold=True,halign='right')
    row=14
    for plan,datos in planes_activos.items():
        c(ws,f'B{row}',plan); ws[f'C{row}'].value=datos['recetas']; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(size=10); row+=1
    c(ws,f'B{row}','TOTAL',bold=True); ws[f'C{row}'].value=total_recetas; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(bold=True,size=10)
    row_end_car=row; box(ws,11,2,row_end_car,3)

    row+=1; ws.row_dimensions[row].height=8
    row_start_pagos=row+1
    ws.merge_cells(f'B{row_start_pagos}:C{row_start_pagos}'); c(ws,f'B{row_start_pagos}','PAGOS A FARMACIA',bold=True,size=11,halign='center'); row=row_start_pagos+1
    ws.merge_cells(f'B{row}:C{row}'); c(ws,f'B{row}','ANTICIPO (OPF)',bold=True,size=10,fill=LIGHT_BLUE,halign='center'); row+=1
    c(ws,f'B{row}','TOTAL',bold=True); n(ws,f'C{row}',opf['efvo_ioma']); row+=1
    c(ws,f'B{row}','Fecha pago'); d(ws,f'C{row}',parse_date(opf['fecha_opf'])); row+=1
    c(ws,f'B{row}','Comprobante'); ws[f'C{row}'].value=opf['nro_comprobante_opf']; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(size=10); row+=1
    ws.merge_cells(f'B{row}:C{row}'); c(ws,f'B{row}','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center'); row+=1
    c(ws,f'B{row}','Bruto a pagar antes imp.:'); n(ws,f'C{row}',pre['total_liquidacion']); row+=1
    c(ws,f'B{row}','TOTAL',bold=True); n(ws,f'C{row}',liq_final); row+=1
    c(ws,f'B{row}','Fecha pago'); d(ws,f'C{row}',parse_date(pago['fecha_pago'])); row+=1
    c(ws,f'B{row}','Comprobante'); ws[f'C{row}'].value=pago['nro_comprobante_pago']; ws[f'C{row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'C{row}'].font=Font(size=10)
    row_end_pagos=row; box(ws,row_start_pagos,2,row_end_pagos,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','IMPUESTOS',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Ing. Btos + Gcias Anticipo:'); n(ws,'F13',opf['ing_brutos_anticipo'])
    c(ws,'E14','Ing. Btos + Gcias Pago Final:'); n(ws,'F14',pago['ing_brutos_pago'])
    c(ws,'E15','TOTAL',bold=True); n(ws,'F15',total_ing_brutos)
    ws.merge_cells('E16:F16'); c(ws,'E16','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E17','TOTAL',bold=True); n(ws,'F17',abs(pre['bonificaciones']))
    ws.merge_cells('E18:F18'); c(ws,'E18','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E19','Colegio Art. 12 SU:'); n(ws,'F19',abs(pre['retencion_colegio_art12']))
    c(ws,'E20','Fdo Prest. COLFARMA:'); n(ws,'F20',abs(pre['fdo_prest_colfarma']))
    c(ws,'E21','TOTAL',bold=True); n(ws,'F21',ret_cofa)
    ws.merge_cells('E22:F22'); c(ws,'E22','DÉB. / CRÉD. OS',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E23','Débito OS:'); n(ws,'F23',deb_os); c(ws,'E24','Crédito OS:'); n(ws,'F24',cred_os)
    box(ws,11,5,24,6)

    ws.merge_cells('H12:L12'); c(ws,'H12','NOTAS DE CRÉDITO',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'I13','Monto Calculado',bold=True,halign='center'); c(ws,'K13','Monto según NR',bold=True,halign='center'); c(ws,'L13','Diferencia',bold=True,halign='center')
    c(ws,'H14','NRF Anticipo:'); n(ws,'I14',abs(pre['nrf_ant']))
    c(ws,'H15','NRF Definitivo:'); n(ws,'I15',abs(pre['nrf_def']))
    c(ws,'H16','NRF Directas:'); n(ws,'I16',abs(pre['nrf_directas']))
    c(ws,'H17','TOTAL',bold=True); n(ws,'I17',notas_cred); n(ws,'K17',total_nr); n(ws,'L17',dif_nr)
    nr_row=19
    ws.merge_cells(f'H{nr_row}:L{nr_row}'); c(ws,f'H{nr_row}','DESGLOSE DE PAGOS',bold=True,size=10,fill=LIGHT_BLUE,halign='center'); nr_row+=1
    c(ws,f'H{nr_row}','Fecha',bold=True,halign='center'); c(ws,f'I{nr_row}','Monto',bold=True,halign='center'); c(ws,f'K{nr_row}','Días',bold=True,halign='center'); nr_row+=1
    for item in sorted(nr_por_fecha, key=lambda x: x.get('fecha','')):
        df=days_diff(fecha_pres,parse_date(item.get('fecha',''))); d(ws,f'H{nr_row}',parse_date(item.get('fecha',''))); n(ws,f'I{nr_row}',item.get('monto',0))
        ws[f'K{nr_row}'].value=df; ws[f'K{nr_row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'K{nr_row}'].font=Font(size=10); nr_row+=1
    c(ws,f'H{nr_row}','DÍAS PROM.',bold=True); ws[f'K{nr_row}'].value=round(dias_nr_pond,1); ws[f'K{nr_row}'].alignment=Alignment(horizontal='right',vertical='center'); ws[f'K{nr_row}'].font=Font(bold=True,size=10)
    box(ws,11,8,nr_row,12)

    ws2=wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines=False
    style=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    h1=['RECETAS','IMPORTE 100%','A/C INSTITUTO','AFILIADO','%AFL','IOMA','%IOMA','DIAS PROM. PAGO','DEBITOS','RETENCIONES COFA','%RET','BONIFICACIONES','%BON','ING BRUTOS RETENIDOS']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[total_recetas,total_importe100,total_ac,afiliado,afiliado/total_importe100 if total_importe100 else 0,total_pagado,total_pagado/total_ac if total_ac else 0,round(dias_prom,2),deb_os,ret_cofa,ret_cofa/total_ac if total_ac else 0,abs(pre['bonificaciones']),abs(pre['bonificaciones'])/total_ac if total_ac else 0,total_ing_brutos]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v; cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_reporte',ref='A1:N2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)
    h2=['ANTICIPO','%IOMA ANT','DIAS ANT','LIQ FINAL','%IOMA LIQ FINAL','DIAS LIQ FINAL','NOTAS DE RECUPERO','%IOMA NR','DIAS NR']
    for i,h in enumerate(h2): ws2.cell(4,i+1,h)
    row5=[opf['efvo_ioma'],opf['efvo_ioma']/total_pagado if total_pagado else 0,dias_ant,liq_final,liq_final/total_pagado if total_pagado else 0,dias_liq,total_nr,total_nr/total_pagado if total_pagado else 0,round(dias_nr_pond,2)]
    for i,v in enumerate(row5):
        cell=ws2.cell(5,i+1); cell.value=v; cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl2=Table(displayName='tbl_desglose',ref='A4:I5'); tbl2.tableStyleInfo=style; ws2.add_table(tbl2)
    ws2.cell(7,1,'Diferencias NR'); cell=ws2.cell(8,1); cell.value=dif_nr; cell.number_format='#,##0.00'
    tbl3=Table(displayName='tbl_diferencias',ref='A7:A8'); tbl3.tableStyleInfo=style; ws2.add_table(tbl3)
    for col in 'ABCDEFGHIJKLMN': ws2.column_dimensions[col].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── OSDE ──────────────────────────────────────────────────────────────────────

def build_osde_excel(data, mes, anio):
    car=data['caratula']; pre=data['pre']; pago=data['pago']; nr=data['nr']
    fecha_pres=parse_date(car['fecha_cierre'])
    dias_pago=days_diff(fecha_pres,parse_date(pago['fecha_pago']))
    dias_nr=days_diff(fecha_pres,parse_date(nr['nr_fecha']))
    total_ret=pre['retencion_fdo_res']+pre['ret_col_art12']
    total_pagado=pre['neto_cobrar']+nr['nr_monto']
    dias_prom=(pre['neto_cobrar']*dias_pago+nr['nr_monto']*dias_nr)/total_pagado if total_pagado else 0
    dif_nr=nr['nr_monto']-pre['notas_credito']
    periodo=f'{mes}/{anio}'

    wb=Workbook(); ws=wb.active; ws.title='Reporte'
    setup_ws(ws); header_bg(ws)

    ws.merge_cells('B2:C3'); c(ws,'B2','OSDE',bold=True,size=28,color=WHITE,fill=DARK_BLUE,halign='center')
    ws.merge_cells('E2:F3'); c(ws,'E2',periodo,bold=True,size=20,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'H2','RECETAS',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'H3',car['nro_recetas'],bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0')
    c(ws,'I2','FECHA DE PRESENTACION',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'I3',fecha_pres.strftime('%d/%m/%Y'),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center')
    c(ws,'K2','DÍAS PROM.',size=9,color='D6E4F0',fill=DARK_BLUE,halign='center')
    c(ws,'K3',round(dias_prom,1),bold=True,size=14,color=WHITE,fill=DARK_BLUE,halign='center',num_fmt='#,##0.0')

    ws.merge_cells('B5:C5'); c(ws,'B5','IMPORTE TOTAL',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('B6:C6'); ni(ws,'B6',car['importe_total']); ws['B6'].font=Font(bold=True,size=13,color=WHITE); ws['B6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['B6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('E5:F5'); c(ws,'E5','A/C OSDE',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('E6:F6'); ni(ws,'E6',car['a_cargo_osde']); ws['E6'].font=Font(bold=True,size=13,color=WHITE); ws['E6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['E6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('H5:I5'); c(ws,'H5','AFILIADO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
    ws.merge_cells('H6:I6'); n(ws,'H6',car['afiliado']); ws['H6'].font=Font(bold=True,size=13,color=WHITE); ws['H6'].fill=PatternFill('solid',fgColor=MID_BLUE); ws['H6'].alignment=Alignment(horizontal='center',vertical='center')
    ws.merge_cells('K5:L5'); c(ws,'K5','TOTAL PAGADO OSDE',size=9,color=WHITE,fill=GREEN,halign='center')
    ws.merge_cells('K6:L6'); ni(ws,'K6',total_pagado); ws['K6'].font=Font(bold=True,size=13,color=WHITE); ws['K6'].fill=PatternFill('solid',fgColor=GREEN); ws['K6'].alignment=Alignment(horizontal='center',vertical='center')

    for col,col_dias,lbl,dias,val in [('B','C','LIQ. FINAL',dias_pago,pre['neto_cobrar']),('E','F','NOTAS RECUP.',dias_nr,nr['nr_monto'])]:
        c(ws,f'{col}8',lbl,size=9,color=WHITE,fill=MID_BLUE,halign='center')
        c(ws,f'{col_dias}8','DIAS DE PAGO',size=9,color=WHITE,fill=MID_BLUE,halign='center')
        n(ws,f'{col}9',val); ws[f'{col}9'].font=Font(bold=True,size=13,color=WHITE); ws[f'{col}9'].fill=PatternFill('solid',fgColor=MID_BLUE); ws[f'{col}9'].alignment=Alignment(horizontal='center',vertical='center')
        c(ws,f'{col_dias}9',dias,bold=True,size=13,color=WHITE,fill=MID_BLUE,halign='center')

    ws.merge_cells('B11:C11'); c(ws,'B11','PAGOS A FARMACIA',bold=True,size=11,halign='center')
    ws.merge_cells('E11:F11'); c(ws,'E11','DESCUENTOS',bold=True,size=11,halign='center')
    ws.merge_cells('H11:L11'); c(ws,'H11','PAGOS A DROGUERIA',bold=True,size=11,halign='center')

    ws.merge_cells('B12:C12'); c(ws,'B12','LIQUIDACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'B13','Total a verificar:'); n(ws,'C13',car['total_verificar'])
    c(ws,'B14','NETO A COBRAR',bold=True); n(ws,'C14',pre['neto_cobrar'])
    c(ws,'B15','Fecha pago'); d(ws,'C15',parse_date(pago['fecha_pago']))
    c(ws,'B16','Comprobante'); ws['C16'].value=pago['nro_comprobante_pago']; ws['C16'].alignment=Alignment(horizontal='right',vertical='center'); ws['C16'].font=Font(size=10)
    box(ws,11,2,16,3)

    ws.merge_cells('E12:F12'); c(ws,'E12','BONIFICACIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E13','Total:'); n(ws,'F13',car['bonificacion'])
    ws.merge_cells('E14:F14'); c(ws,'E14','RETENCIONES',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E15','Fdo. Res.:'); n(ws,'F15',pre['retencion_fdo_res'])
    c(ws,'E16','Colegio Art. 12 SU:'); n(ws,'F16',pre['ret_col_art12'])
    c(ws,'E17','TOTAL',bold=True); n(ws,'F17',total_ret)
    ws.merge_cells('E18:F18'); c(ws,'E18','AJUSTE FACTURACIÓN',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'E19','Débito:'); n(ws,'F19',pre['ajuste_facturacion'])
    box(ws,11,5,19,6)

    ws.merge_cells('H12:L12'); c(ws,'H12','NOTAS DE CRÉDITO',bold=True,size=10,fill=LIGHT_BLUE,halign='center')
    c(ws,'I13','Monto Calculado',bold=True,halign='center'); c(ws,'K13','Monto según NR',bold=True,halign='center'); c(ws,'L13','Diferencia',bold=True,halign='center')
    c(ws,'H14','TOTAL',bold=True); n(ws,'I14',pre['notas_credito']); n(ws,'K14',nr['nr_monto']); n(ws,'L14',dif_nr)
    c(ws,'H15','Fecha NR:'); d(ws,'I15',parse_date(nr['nr_fecha']))
    c(ws,'H16','Días:'); ws['I16'].value=dias_nr; ws['I16'].alignment=Alignment(horizontal='right',vertical='center'); ws['I16'].font=Font(size=10)
    box(ws,11,8,16,12)

    ws2=wb.create_sheet('Resumen'); ws2.sheet_view.showGridLines=False
    style=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    h1=['RECETAS','IMPORTE TOTAL','A/C OSDE','AFILIADO','%AFL','TOTAL PAGADO','%PAGADO','DIAS PROM. PAGO','AJUSTE FACTURACION','RETENCIONES','%RET','BONIFICACIONES','%BON']
    for i,h in enumerate(h1): ws2.cell(1,i+1,h)
    row2=[car['nro_recetas'],car['importe_total'],car['a_cargo_osde'],car['afiliado'],car['afiliado']/car['importe_total'] if car['importe_total'] else 0,total_pagado,total_pagado/car['a_cargo_osde'] if car['a_cargo_osde'] else 0,round(dias_prom,2),pre['ajuste_facturacion'],total_ret,total_ret/car['a_cargo_osde'] if car['a_cargo_osde'] else 0,car['bonificacion'],car['bonificacion']/car['a_cargo_osde'] if car['a_cargo_osde'] else 0]
    for i,v in enumerate(row2):
        cell=ws2.cell(2,i+1); cell.value=v; cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl1=Table(displayName='tbl_reporte',ref='A1:M2'); tbl1.tableStyleInfo=style; ws2.add_table(tbl1)
    h2=['LIQ FINAL','%OSDE LIQ','DIAS LIQ','NOTAS RECUPERO','%OSDE NR','DIAS NR']
    for i,h in enumerate(h2): ws2.cell(4,i+1,h)
    row5=[pre['neto_cobrar'],pre['neto_cobrar']/total_pagado if total_pagado else 0,dias_pago,nr['nr_monto'],nr['nr_monto']/total_pagado if total_pagado else 0,dias_nr]
    for i,v in enumerate(row5):
        cell=ws2.cell(5,i+1); cell.value=v; cell.number_format='0.00%' if isinstance(v,float) and abs(v)<10 else '#,##0.00'
    tbl2=Table(displayName='tbl_desglose',ref='A4:F5'); tbl2.tableStyleInfo=style; ws2.add_table(tbl2)
    ws2.cell(7,1,'Diferencias NR'); cell=ws2.cell(8,1); cell.value=dif_nr; cell.number_format='#,##0.00'
    tbl3=Table(displayName='tbl_diferencias',ref='A7:A8'); tbl3.tableStyleInfo=style; ws2.add_table(tbl3)
    for col in 'ABCDEFGHIJKLM': ws2.column_dimensions[col].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── ENDPOINTS ─────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "service": "Farmacia Merlo - Generador de Reportes"}

# Importar routers
from reportes import router as reportes_router
from debitos import router as debitos_router

app.include_router(reportes_router)
app.include_router(debitos_router)
